import calendar
import csv
import io
import os
import sqlite3
import uuid
import zipfile
from collections.abc import Mapping
from datetime import date, datetime
from functools import wraps

from flask import Flask, Response, flash, jsonify, redirect, render_template, request, session, url_for, send_file, send_from_directory, has_request_context
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
except Exception:  # pragma: no cover
    SimpleDocTemplate = None

APP_NAME = "Gestão360 Contábil"
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
IS_VERCEL = bool(os.environ.get("VERCEL"))
RUNTIME_DIR = os.environ.get("GESTAO360_DATA_DIR") or (os.path.join("/tmp", "gestao360") if IS_VERCEL else BASE_DIR)
DB_PATH = os.path.join(RUNTIME_DIR, "gestao360.db")
UPLOAD_DIR = os.path.join(RUNTIME_DIR, "uploads", "activity_files")
BACKUP_DIR = os.path.join(RUNTIME_DIR, "backups")
TURSO_DATABASE_URL = (os.environ.get("TURSO_DATABASE_URL") or "").strip()
TURSO_AUTH_TOKEN = (os.environ.get("TURSO_AUTH_TOKEN") or "").strip()

app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.environ.get("GESTAO360_SECRET", "desenvolvimento-local-troque-em-producao"),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=IS_VERCEL or os.environ.get("GESTAO360_HTTPS", "0") == "1",
    MAX_CONTENT_LENGTH=20 * 1024 * 1024,
)

CLIENT_TYPES = ["Pessoa Física", "Pessoa Jurídica", "Produtor Rural", "Outro"]
CLIENT_STATUS = ["Ativo", "Inativo", "Em análise"]
SERVICE_TYPES = [
    "Folha de pagamento", "Fiscal", "Contábil", "IRPF", "Produtor Rural",
    "Regularização", "Atendimento", "Reunião", "Documento", "Outro"
]
ACTIVITY_STATUS = ["Pendente", "Em andamento", "Aguardando cliente", "Aguardando documentos", "Em revisão", "Concluído"]
ACTIVITY_PRIORITIES = ["Baixa", "Média", "Alta", "Urgente"]
RECEIVABLE_STATUS = ["A receber", "Parcial", "Parcial em atraso", "Pago", "Atrasado", "Cancelado"]
PAYMENT_METHODS = ["Pix", "Boleto", "Dinheiro", "Cartão", "Transferência", "Outro"]


class CompatRow(Mapping):
    """Linha compatível com sqlite3.Row para conexões libSQL remotas."""

    def __init__(self, columns, values):
        self._columns = tuple(columns)
        self._values = tuple(values)
        self._data = dict(zip(self._columns, self._values))

    def __getitem__(self, key):
        if isinstance(key, int):
            return self._values[key]
        return self._data[key]

    def __iter__(self):
        return iter(self._columns)

    def __len__(self):
        return len(self._columns)

    def keys(self):
        return self._columns

    def __getattr__(self, key):
        try:
            return self._data[key]
        except KeyError as exc:
            raise AttributeError(key) from exc


class CompatCursor:
    def __init__(self, cursor):
        self._cursor = cursor
        self._columns = tuple(item[0] for item in ((cursor.description if cursor else ()) or ()))

    def _wrap(self, row):
        if row is None or not self._columns:
            return row
        return CompatRow(self._columns, row)

    def fetchone(self):
        return self._wrap(self._cursor.fetchone())

    def fetchall(self):
        return [self._wrap(row) for row in self._cursor.fetchall()]

    def __iter__(self):
        while True:
            row = self.fetchone()
            if row is None:
                return
            yield row

    @property
    def lastrowid(self):
        return self._cursor.lastrowid

    @property
    def rowcount(self):
        return getattr(self._cursor, "rowcount", -1)

    @property
    def description(self):
        return self._cursor.description

    def __getattr__(self, name):
        return getattr(self._cursor, name)


class CompatConnection:
    def __init__(self, connection):
        self._connection = connection

    def execute(self, sql, params=()):
        return CompatCursor(self._connection.execute(sql, params))

    def executemany(self, sql, params):
        return CompatCursor(self._connection.executemany(sql, params))

    def executescript(self, script):
        cursor = self._connection.executescript(script)
        return CompatCursor(cursor) if cursor is not None else None

    def commit(self):
        return self._connection.commit()

    def rollback(self):
        return self._connection.rollback()

    def close(self):
        return self._connection.close()

    def __enter__(self):
        self._connection.__enter__()
        return self

    def __exit__(self, exc_type, exc, tb):
        return self._connection.__exit__(exc_type, exc, tb)

    def __getattr__(self, name):
        return getattr(self._connection, name)


def using_remote_database():
    return bool(TURSO_DATABASE_URL)


def get_db():
    if using_remote_database():
        try:
            import libsql
        except ImportError as exc:  # pragma: no cover - depende do ambiente de produção
            raise RuntimeError("Instale a dependência libsql para usar o banco Turso.") from exc
        connect_kwargs = {"database": TURSO_DATABASE_URL}
        if TURSO_AUTH_TOKEN:
            connect_kwargs["auth_token"] = TURSO_AUTH_TOKEN
        conn = libsql.connect(**connect_kwargs)
        conn.execute("PRAGMA foreign_keys = ON")
        return CompatConnection(conn)

    os.makedirs(RUNTIME_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def query_all(sql, params=()):
    with get_db() as conn:
        return conn.execute(sql, params).fetchall()


def query_one(sql, params=()):
    with get_db() as conn:
        return conn.execute(sql, params).fetchone()


def execute(sql, params=()):
    with get_db() as conn:
        cur = conn.execute(sql, params)
        conn.commit()
        return cur.lastrowid


def current_audit_context():
    if has_request_context():
        return {
            "user_id": session.get("user_id"),
            "user_name": session.get("name") or "Sistema",
            "ip_address": request.headers.get("X-Forwarded-For", request.remote_addr or ""),
        }
    return {"user_id": None, "user_name": "Sistema", "ip_address": ""}


def add_audit_log_conn(conn, action, entity_type, entity_id=None, entity_label="", details=""):
    ctx = current_audit_context()
    conn.execute(
        """
        INSERT INTO audit_logs
            (user_id, user_name, action, entity_type, entity_id, entity_label, details, ip_address)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            ctx["user_id"],
            ctx["user_name"],
            action,
            entity_type,
            entity_id,
            (entity_label or "")[:255],
            (details or "")[:1200],
            ctx["ip_address"],
        ),
    )


def audit_log(action, entity_type, entity_id=None, entity_label="", details=""):
    try:
        with get_db() as conn:
            add_audit_log_conn(conn, action, entity_type, entity_id, entity_label, details)
            conn.commit()
    except Exception:
        # O histórico não pode impedir a operação principal, principalmente durante migrações.
        pass


def add_months_to_date(date_string, months_to_add):
    base = datetime.strptime(date_string, "%Y-%m-%d").date()
    month_index = (base.month - 1) + months_to_add
    year = base.year + month_index // 12
    month = (month_index % 12) + 1
    day = min(base.day, calendar.monthrange(year, month)[1])
    return date(year, month, day).isoformat()


def date_for_month_day(month_string, day_value):
    if not month_string:
        month_string = current_month()
    year, mon = map(int, month_string.split("-"))
    try:
        day = int(day_value or 1)
    except (TypeError, ValueError):
        day = 1
    max_day = calendar.monthrange(year, mon)[1]
    day = min(max(day, 1), max_day)
    return date(year, mon, day).isoformat()


def init_db():
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)
    with get_db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('Administrador', 'Colaborador')),
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                document TEXT,
                client_type TEXT,
                phone TEXT,
                email TEXT,
                responsible TEXT,
                monthly_fee REAL NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'Ativo',
                notes TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS activities (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_id INTEGER NOT NULL,
                activity_date TEXT NOT NULL,
                service_type TEXT NOT NULL,
                description TEXT NOT NULL,
                collaborator TEXT,
                time_spent REAL NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'Concluído',
                due_date TEXT,
                priority TEXT NOT NULL DEFAULT 'Média',
                notes TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT,
                recurrence_template_id INTEGER,
                recurrence_month TEXT,
                FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS receivables (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_id INTEGER NOT NULL,
                description TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                paid_amount REAL NOT NULL DEFAULT 0,
                due_date TEXT NOT NULL,
                payment_date TEXT,
                status TEXT NOT NULL DEFAULT 'A receber',
                payment_method TEXT,
                notes TEXT,
                installment_number INTEGER NOT NULL DEFAULT 1,
                installment_total INTEGER NOT NULL DEFAULT 1,
                receivable_group_id TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS receivable_payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receivable_id INTEGER NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                payment_date TEXT NOT NULL,
                payment_method TEXT,
                notes TEXT,
                created_by TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(receivable_id) REFERENCES receivables(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS activity_time_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                activity_id INTEGER NOT NULL,
                user_id INTEGER,
                user_name TEXT,
                started_at TEXT,
                ended_at TEXT,
                seconds INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(activity_id) REFERENCES activities(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS activity_updates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                activity_id INTEGER NOT NULL,
                user_id INTEGER,
                user_name TEXT,
                message TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(activity_id) REFERENCES activities(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS activity_checklist_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                activity_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                is_done INTEGER NOT NULL DEFAULT 0,
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_by TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                completed_at TEXT,
                completed_by TEXT,
                FOREIGN KEY(activity_id) REFERENCES activities(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS checklist_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                service_type TEXT,
                active INTEGER NOT NULL DEFAULT 1,
                created_by TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS checklist_template_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(template_id) REFERENCES checklist_templates(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS recurring_activity_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_id INTEGER,
                service_type TEXT NOT NULL,
                description TEXT NOT NULL,
                collaborator TEXT,
                priority TEXT NOT NULL DEFAULT 'Média',
                activity_day INTEGER NOT NULL DEFAULT 1,
                due_day INTEGER,
                checklist_template_id INTEGER,
                active INTEGER NOT NULL DEFAULT 1,
                notes TEXT,
                created_by TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT,
                FOREIGN KEY(client_id) REFERENCES clients(id) ON DELETE CASCADE,
                FOREIGN KEY(checklist_template_id) REFERENCES checklist_templates(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS activity_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                activity_id INTEGER NOT NULL,
                original_filename TEXT NOT NULL,
                stored_filename TEXT,
                uploaded_by TEXT,
                size_bytes INTEGER DEFAULT 0,
                mime_type TEXT,
                file_content BLOB,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(activity_id) REFERENCES activities(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                user_name TEXT,
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                entity_id INTEGER,
                entity_label TEXT,
                details TEXT,
                ip_address TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            """
        )

        activity_columns = {row[1] for row in conn.execute("PRAGMA table_info(activities)").fetchall()}
        if "due_date" not in activity_columns:
            conn.execute("ALTER TABLE activities ADD COLUMN due_date TEXT")
        if "priority" not in activity_columns:
            conn.execute("ALTER TABLE activities ADD COLUMN priority TEXT NOT NULL DEFAULT 'Média'")
        if "updated_at" not in activity_columns:
            conn.execute("ALTER TABLE activities ADD COLUMN updated_at TEXT")
        if "recurrence_template_id" not in activity_columns:
            conn.execute("ALTER TABLE activities ADD COLUMN recurrence_template_id INTEGER")
        if "recurrence_month" not in activity_columns:
            conn.execute("ALTER TABLE activities ADD COLUMN recurrence_month TEXT")

        recurring_columns = {row[1] for row in conn.execute("PRAGMA table_info(recurring_activity_templates)").fetchall()}
        if recurring_columns:
            if "checklist_template_id" not in recurring_columns:
                conn.execute("ALTER TABLE recurring_activity_templates ADD COLUMN checklist_template_id INTEGER")
            if "notes" not in recurring_columns:
                conn.execute("ALTER TABLE recurring_activity_templates ADD COLUMN notes TEXT")
            if "updated_at" not in recurring_columns:
                conn.execute("ALTER TABLE recurring_activity_templates ADD COLUMN updated_at TEXT")

        receivable_columns = {row[1] for row in conn.execute("PRAGMA table_info(receivables)").fetchall()}
        if "installment_number" not in receivable_columns:
            conn.execute("ALTER TABLE receivables ADD COLUMN installment_number INTEGER NOT NULL DEFAULT 1")
        if "installment_total" not in receivable_columns:
            conn.execute("ALTER TABLE receivables ADD COLUMN installment_total INTEGER NOT NULL DEFAULT 1")
        if "receivable_group_id" not in receivable_columns:
            conn.execute("ALTER TABLE receivables ADD COLUMN receivable_group_id TEXT")
        if "paid_amount" not in receivable_columns:
            conn.execute("ALTER TABLE receivables ADD COLUMN paid_amount REAL NOT NULL DEFAULT 0")
        if "cancelled_at" not in receivable_columns:
            conn.execute("ALTER TABLE receivables ADD COLUMN cancelled_at TEXT")
        if "cancelled_by" not in receivable_columns:
            conn.execute("ALTER TABLE receivables ADD COLUMN cancelled_by TEXT")
        if "updated_at" not in receivable_columns:
            conn.execute("ALTER TABLE receivables ADD COLUMN updated_at TEXT")

        file_columns = {row[1] for row in conn.execute("PRAGMA table_info(activity_files)").fetchall()}
        if file_columns:
            if "mime_type" not in file_columns:
                conn.execute("ALTER TABLE activity_files ADD COLUMN mime_type TEXT")
            if "file_content" not in file_columns:
                conn.execute("ALTER TABLE activity_files ADD COLUMN file_content BLOB")

        # Compatibilidade com cobranças antigas: parcelas já pagas passam a ter valor baixado registrado.
        conn.execute("""
            UPDATE receivables
               SET paid_amount = amount
             WHERE status = 'Pago'
               AND COALESCE(paid_amount, 0) <= 0
        """)
        conn.execute("""
            INSERT INTO receivable_payments (receivable_id, amount, payment_date, payment_method, notes, created_by)
            SELECT r.id,
                   r.amount,
                   COALESCE(r.payment_date, DATE(r.created_at), DATE('now')),
                   r.payment_method,
                   'Migração automática de baixa já existente',
                   'Sistema'
              FROM receivables r
             WHERE r.status = 'Pago'
               AND COALESCE(r.amount, 0) > 0
               AND NOT EXISTS (SELECT 1 FROM receivable_payments p WHERE p.receivable_id = r.id)
        """)

        # Remove somente o usuário de demonstração que ainda estiver com a senha original.
        # Contas que já tiveram a senha alterada são preservadas.
        demo_user = conn.execute("SELECT id, password_hash FROM users WHERE email = ?", ("admin@gestao360.com",)).fetchone()
        if demo_user and check_password_hash(demo_user["password_hash"], "admin123"):
            conn.execute("DELETE FROM users WHERE id = ?", (demo_user["id"],))

        if not conn.execute("SELECT key FROM settings WHERE key = 'hide_financial_for_collab'").fetchone():
            conn.execute("INSERT INTO settings (key, value) VALUES ('hide_financial_for_collab', '1')")
        if not conn.execute("SELECT key FROM settings WHERE key = 'allow_self_registration'").fetchone():
            conn.execute("INSERT INTO settings (key, value) VALUES ('allow_self_registration', '1')")
        if not conn.execute("SELECT key FROM settings WHERE key = 'registration_code_hash'").fetchone():
            conn.execute("INSERT INTO settings (key, value) VALUES ('registration_code_hash', '')")
        conn.commit()


def money(value):
    try:
        value = float(value or 0)
    except (TypeError, ValueError):
        value = 0
    formatted = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def parse_money(value):
    if value is None:
        return 0.0
    value = str(value).strip()
    value = value.replace("R$", "").replace(" ", "")
    if "," in value:
        value = value.replace(".", "").replace(",", ".")
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def hours(value):
    try:
        value = float(value or 0)
    except (TypeError, ValueError):
        value = 0
    return f"{value:.2f} h".replace(".", ",")


def month_label(month):
    try:
        dt = datetime.strptime(month + "-01", "%Y-%m-%d")
        return dt.strftime("%m/%Y")
    except Exception:
        return month


def current_month():
    return date.today().strftime("%Y-%m")


def setting_value(key, default=""):
    row = query_one("SELECT value FROM settings WHERE key = ?", (key,))
    return row["value"] if row else default


def system_has_users():
    row = query_one("SELECT COUNT(*) AS total FROM users")
    return bool(row and int(row["total"] or 0) > 0)


def validate_password(password):
    if len(password or "") < 8:
        return "A senha precisa ter pelo menos 8 caracteres."
    if (password or "").lower() in {"admin123", "12345678", "senha123", "gestao360", "password"}:
        return "Escolha uma senha diferente das senhas comuns."
    if not any(ch.isalpha() for ch in password) or not any(ch.isdigit() for ch in password):
        return "Use pelo menos uma letra e um número na senha."
    return None


def login_user_session(user):
    session.clear()
    session["user_id"] = user["id"]
    session["name"] = user["name"]
    session["email"] = user["email"]
    session["role"] = user["role"]
    session["using_default_password"] = False


def can_view_financial():
    if session.get("role") == "Administrador":
        return True
    hidden = query_one("SELECT value FROM settings WHERE key = 'hide_financial_for_collab'")
    return not (hidden and hidden["value"] == "1")


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if session.get("role") != "Administrador":
            flash("Acesso restrito ao administrador.", "error")
            return redirect(url_for("dashboard"))
        return fn(*args, **kwargs)
    return wrapper


def financial_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not can_view_financial():
            flash("Seu usuário não tem permissão para ver informações financeiras.", "error")
            return redirect(url_for("dashboard"))
        return fn(*args, **kwargs)
    return wrapper


def month_range(month):
    if not month:
        month = current_month()
    start = month + "-01"
    year, mon = map(int, month.split("-"))
    if mon == 12:
        end = f"{year + 1}-01-01"
    else:
        end = f"{year}-{mon + 1:02d}-01"
    return start, end


def row_get(row, key, default=None):
    try:
        return row[key]
    except (KeyError, IndexError, TypeError):
        return default


def receivable_base_amount(row):
    try:
        return float(row_get(row, "amount", 0) or 0)
    except (TypeError, ValueError):
        return 0.0


def receivable_paid_amount(row):
    try:
        paid = float(row_get(row, "paid_amount", 0) or 0)
    except (TypeError, ValueError):
        paid = 0.0
    # Compatibilidade com bases antigas que tinham apenas status Pago.
    if paid <= 0 and row_get(row, "status") == "Pago":
        paid = receivable_base_amount(row)
    amount = receivable_base_amount(row)
    if amount > 0:
        paid = min(paid, amount)
    return max(paid, 0.0)


def receivable_remaining_amount(row):
    if row_get(row, "status") == "Cancelado":
        return 0.0
    return max(receivable_base_amount(row) - receivable_paid_amount(row), 0.0)


def has_receivable_payment(row):
    return receivable_paid_amount(row) > 0.009


def is_receivable_overdue(row):
    return (
        row_get(row, "status") != "Cancelado"
        and receivable_remaining_amount(row) > 0.009
        and bool(row_get(row, "due_date"))
        and row_get(row, "due_date") < date.today().isoformat()
    )


def computed_receivable_status(row):
    status = row_get(row, "status")
    if status == "Cancelado":
        return "Cancelado"

    amount = receivable_base_amount(row)
    paid = receivable_paid_amount(row)
    remaining = max(amount - paid, 0.0)

    if amount > 0 and remaining <= 0.009:
        return "Pago"
    if paid > 0.009:
        return "Parcial em atraso" if is_receivable_overdue(row) else "Parcial"
    if is_receivable_overdue(row):
        return "Atrasado"
    return "A receber"


def is_receivable_paid(row):
    return computed_receivable_status(row) == "Pago"


def is_receivable_cancelled(row):
    return computed_receivable_status(row) == "Cancelado"


def can_delete_receivable(row):
    # Cobrança com qualquer baixa deve permanecer no histórico financeiro.
    return not has_receivable_payment(row) and not is_receivable_paid(row)


def can_cancel_receivable(row):
    # Cancelar é para lançamento em aberto/atrasado sem baixa registrada.
    return not has_receivable_payment(row) and not is_receivable_paid(row) and not is_receivable_cancelled(row)


def can_receive_payment(row):
    return not is_receivable_cancelled(row) and receivable_remaining_amount(row) > 0.009


def receivable_amounts_summary(rows):
    active_rows = [r for r in rows if computed_receivable_status(r) != "Cancelado"]
    total_amount = sum(receivable_base_amount(r) for r in active_rows)
    paid_amount = sum(receivable_paid_amount(r) for r in active_rows)
    open_amount = sum(receivable_remaining_amount(r) for r in active_rows)
    cancelled_amount = sum(receivable_base_amount(r) for r in rows if computed_receivable_status(r) == "Cancelado")
    paid_count = sum(1 for r in active_rows if computed_receivable_status(r) == "Pago")
    cancelled_count = sum(1 for r in rows if computed_receivable_status(r) == "Cancelado")
    partial_count = sum(1 for r in active_rows if computed_receivable_status(r) in ("Parcial", "Parcial em atraso"))
    return total_amount, paid_amount, open_amount, cancelled_amount, paid_count, cancelled_count, partial_count


def rentability_label(value_per_hour):
    if value_per_hour is None:
        return "Sem horas"
    if value_per_hour >= 150:
        return "Muito rentável"
    if value_per_hour >= 80:
        return "Rentável"
    if value_per_hour >= 40:
        return "Atenção"
    return "Pouco rentável"


def seconds_to_hours(seconds):
    try:
        return round(int(seconds or 0) / 3600, 4)
    except (TypeError, ValueError):
        return 0


def format_seconds(seconds):
    try:
        seconds = int(seconds or 0)
    except (TypeError, ValueError):
        seconds = 0
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def get_activity_or_404(activity_id):
    return query_one(
        """
        SELECT a.*, c.name AS client_name, rt.description AS recurrence_template_description
        FROM activities a
        JOIN clients c ON c.id = a.client_id
        LEFT JOIN recurring_activity_templates rt ON rt.id = a.recurrence_template_id
        WHERE a.id = ?
        """,
        (activity_id,),
    )


def activity_next_status(current_status, direction="next"):
    if current_status not in ACTIVITY_STATUS:
        return ACTIVITY_STATUS[0]
    index = ACTIVITY_STATUS.index(current_status)
    if direction == "previous":
        return ACTIVITY_STATUS[max(index - 1, 0)]
    return ACTIVITY_STATUS[min(index + 1, len(ACTIVITY_STATUS) - 1)]


def activity_is_overdue(row):
    return (
        row_get(row, "status") != "Concluído"
        and bool(row_get(row, "due_date"))
        and row_get(row, "due_date") < date.today().isoformat()
    )


def activity_priority_order(priority):
    order = {"Urgente": 0, "Alta": 1, "Média": 2, "Baixa": 3}
    return order.get(priority or "Média", 2)


def get_activity_checklist_stats(activity_ids):
    if not activity_ids:
        return {}
    placeholders = ",".join("?" for _ in activity_ids)
    rows = query_all(
        f"""
        SELECT activity_id,
               COUNT(*) AS total,
               COALESCE(SUM(CASE WHEN is_done = 1 THEN 1 ELSE 0 END), 0) AS done
        FROM activity_checklist_items
        WHERE activity_id IN ({placeholders})
        GROUP BY activity_id
        """,
        activity_ids,
    )
    stats = {int(row["activity_id"]): {"total": int(row["total"] or 0), "done": int(row["done"] or 0), "percent": 0} for row in rows}
    for value in stats.values():
        value["percent"] = int(round((value["done"] / value["total"]) * 100)) if value["total"] else 0
    return stats


def checklist_progress(items):
    total = len(items or [])
    done = sum(1 for item in (items or []) if int(row_get(item, "is_done", 0) or 0) == 1)
    percent = int(round((done / total) * 100)) if total else 0
    return {"total": total, "done": done, "percent": percent}


def parse_checklist_lines(raw_text):
    lines = []
    seen = set()
    for line in (raw_text or "").splitlines():
        title = line.strip().lstrip("-•*0123456789. )").strip()
        normalized = title.casefold()
        if title and normalized not in seen:
            seen.add(normalized)
            lines.append(title)
    return lines


def get_active_checklist_templates():
    return query_all(
        """
        SELECT t.*,
               COUNT(i.id) AS items_count
        FROM checklist_templates t
        LEFT JOIN checklist_template_items i ON i.template_id = t.id
        WHERE t.active = 1
        GROUP BY t.id
        ORDER BY t.name COLLATE NOCASE
        """
    )


def apply_checklist_template_to_activity(conn, activity_id, checklist_template_id, created_by=None):
    if not checklist_template_id:
        return 0
    items = conn.execute(
        """
        SELECT title
        FROM checklist_template_items
        WHERE template_id = ?
        ORDER BY sort_order ASC, id ASC
        """,
        (checklist_template_id,),
    ).fetchall()
    if not items:
        return 0
    existing_rows = conn.execute(
        "SELECT title FROM activity_checklist_items WHERE activity_id = ?",
        (activity_id,),
    ).fetchall()
    existing = {row["title"].strip().casefold() for row in existing_rows}
    max_order_row = conn.execute(
        "SELECT COALESCE(MAX(sort_order), 0) AS max_order FROM activity_checklist_items WHERE activity_id = ?",
        (activity_id,),
    ).fetchone()
    sort_order = int(max_order_row["max_order"] or 0)
    added = 0
    for item in items:
        title = (item["title"] or "").strip()
        if not title:
            continue
        normalized = title.casefold()
        if normalized in existing:
            continue
        sort_order += 1
        existing.add(normalized)
        conn.execute(
            """
            INSERT INTO activity_checklist_items (activity_id, title, sort_order, created_by)
            VALUES (?, ?, ?, ?)
            """,
            (activity_id, title, sort_order, created_by),
        )
        added += 1
    return added


def generate_recurring_activities_for_month(month):
    month = month or current_month()
    created = 0
    skipped = 0
    generated_rows = []

    with get_db() as conn:
        templates = conn.execute(
            """
            SELECT rt.*, c.name AS client_name, c.status AS client_status, ct.name AS checklist_template_name
            FROM recurring_activity_templates rt
            LEFT JOIN clients c ON c.id = rt.client_id
            LEFT JOIN checklist_templates ct ON ct.id = rt.checklist_template_id
            WHERE rt.active = 1
              AND (rt.client_id IS NULL OR COALESCE(c.status, '') != 'Inativo')
            ORDER BY COALESCE(c.name, 'Todos'), rt.service_type, rt.description
            """
        ).fetchall()

        for template in templates:
            if template["client_id"]:
                clients = conn.execute(
                    "SELECT id, name FROM clients WHERE id = ? AND status != 'Inativo'",
                    (template["client_id"],),
                ).fetchall()
            else:
                clients = conn.execute(
                    "SELECT id, name FROM clients WHERE status = 'Ativo' ORDER BY name COLLATE NOCASE"
                ).fetchall()

            for client in clients:
                exists = conn.execute(
                    """
                    SELECT id
                    FROM activities
                    WHERE recurrence_template_id = ?
                      AND recurrence_month = ?
                      AND client_id = ?
                    LIMIT 1
                    """,
                    (template["id"], month, client["id"]),
                ).fetchone()
                if exists:
                    skipped += 1
                    continue

                activity_date = date_for_month_day(month, template["activity_day"] or 1)
                due_date = date_for_month_day(month, template["due_day"]) if template["due_day"] else None
                activity_id = conn.execute(
                    """
                    INSERT INTO activities
                        (client_id, activity_date, service_type, description, collaborator, time_spent, status, due_date, priority, notes, recurrence_template_id, recurrence_month)
                    VALUES (?, ?, ?, ?, ?, 0, 'Pendente', ?, ?, ?, ?, ?)
                    """,
                    (
                        client["id"],
                        activity_date,
                        template["service_type"],
                        template["description"],
                        template["collaborator"],
                        due_date,
                        template["priority"] or "Média",
                        template["notes"],
                        template["id"],
                        month,
                    ),
                ).lastrowid
                checklist_added = apply_checklist_template_to_activity(
                    conn, activity_id, template["checklist_template_id"], session.get("name")
                )
                conn.execute(
                    """
                    INSERT INTO activity_updates (activity_id, user_id, user_name, message)
                    VALUES (?, ?, ?, ?)
                    """,
                    (
                        activity_id,
                        session.get("user_id"),
                        session.get("name"),
                        f"Atividade gerada automaticamente pela rotina mensal de {month_label(month)}."
                        + (f" Checklist aplicado com {checklist_added} item(ns)." if checklist_added else ""),
                    ),
                )
                generated_rows.append({
                    "activity_id": activity_id,
                    "client_name": client["name"],
                    "description": template["description"],
                    "checklist_added": checklist_added,
                })
                created += 1
        conn.commit()

    return {"created": created, "skipped": skipped, "generated": generated_rows}


def add_activity_history(activity_id, message):
    execute(
        """
        INSERT INTO activity_updates (activity_id, user_id, user_name, message)
        VALUES (?, ?, ?, ?)
        """,
        (activity_id, session.get("user_id"), session.get("name"), message),
    )


def store_uploaded_activity_file(activity_id, file_storage):
    if not file_storage or not file_storage.filename:
        return None
    original = file_storage.filename
    safe_original = secure_filename(original) or "arquivo"
    mime_type = file_storage.mimetype or "application/octet-stream"

    if using_remote_database():
        content = file_storage.read()
        execute(
            """
            INSERT INTO activity_files
                (activity_id, original_filename, stored_filename, uploaded_by, size_bytes, mime_type, file_content)
            VALUES (?, ?, NULL, ?, ?, ?, ?)
            """,
            (activity_id, original, session.get("name"), len(content), mime_type, content),
        )
        return original

    stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    stored = f"atividade_{activity_id}_{stamp}_{safe_original}"
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    destination = os.path.join(UPLOAD_DIR, stored)
    file_storage.save(destination)
    size = os.path.getsize(destination) if os.path.exists(destination) else 0
    execute(
        """
        INSERT INTO activity_files
            (activity_id, original_filename, stored_filename, uploaded_by, size_bytes, mime_type)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (activity_id, original, stored, session.get("name"), size, mime_type),
    )
    return stored


def build_monthly_report(month, client_id=None, collaborator=None, service_type=None):
    start, end = month_range(month)
    filters = ["1=1"]
    params = [start, end, start, end]

    if client_id:
        filters.append("c.id = ?")
        params.append(client_id)

    activity_join_filters = ["a.activity_date >= ?", "a.activity_date < ?"]
    activity_params = [start, end]
    if collaborator:
        activity_join_filters.append("a.collaborator LIKE ?")
        activity_params.append(f"%{collaborator}%")
    if service_type:
        activity_join_filters.append("a.service_type = ?")
        activity_params.append(service_type)

    where_clients = " AND ".join(filters)
    activity_filter_sql = " AND ".join(activity_join_filters)

    sql = f"""
        SELECT
            c.id,
            c.name,
            c.monthly_fee,
            c.status,
            COALESCE(SUM(a.time_spent), 0) AS total_hours,
            COUNT(a.id) AS activities_count,
            COALESCE((
                SELECT SUM(p.amount)
                FROM receivable_payments p
                JOIN receivables r ON r.id = p.receivable_id
                WHERE r.client_id = c.id
                  AND p.payment_date >= ?
                  AND p.payment_date < ?
            ), 0) AS received_amount,
            COALESCE((
                SELECT SUM(MAX(r.amount - COALESCE(r.paid_amount, 0), 0))
                FROM receivables r
                WHERE r.client_id = c.id
                  AND r.status != 'Cancelado'
                  AND r.due_date < DATE('now')
                  AND (r.amount - COALESCE(r.paid_amount, 0)) > 0
            ), 0) AS overdue_amount
        FROM clients c
        LEFT JOIN activities a
          ON a.client_id = c.id
         AND {activity_filter_sql}
        WHERE {where_clients}
        GROUP BY c.id
        ORDER BY total_hours DESC, c.name ASC
    """
    full_params = [start, end] + activity_params + ([] if not client_id else [client_id])
    rows = query_all(sql, full_params)

    result = []
    for row in rows:
        total_hours = float(row["total_hours"] or 0)
        monthly_fee = float(row["monthly_fee"] or 0)
        value_per_hour = None if total_hours <= 0 else monthly_fee / total_hours
        financial_status = "Em dia"
        if float(row["overdue_amount"] or 0) > 0:
            financial_status = "Atrasado"
        elif float(row["received_amount"] or 0) <= 0 and monthly_fee > 0:
            financial_status = "Sem recebimento no mês"
        result.append({
            "id": row["id"],
            "name": row["name"],
            "monthly_fee": monthly_fee,
            "total_hours": total_hours,
            "activities_count": int(row["activities_count"] or 0),
            "received_amount": float(row["received_amount"] or 0),
            "financial_status": financial_status,
            "value_per_hour": value_per_hour,
            "rentability": rentability_label(value_per_hour),
            "score_sort": value_per_hour if value_per_hour is not None else -1,
        })
    return result



def build_receipts_report(month, client_id=None, payment_method=None, created_by=None):
    """Retorna somente as baixas efetivamente registradas no período.

    Uma parcela pode aparecer mais de uma vez quando recebeu baixas parciais em datas
    diferentes. A competência deste relatório é sempre a data da baixa
    (receivable_payments.payment_date), nunca a data de vencimento da cobrança.
    """
    start, end = month_range(month)
    filters = ["p.payment_date >= ?", "p.payment_date < ?"]
    params = [start, end]

    if client_id:
        filters.append("c.id = ?")
        params.append(client_id)
    if payment_method:
        if payment_method == "Não informado":
            filters.append("COALESCE(NULLIF(TRIM(p.payment_method), ''), 'Não informado') = 'Não informado'")
        else:
            filters.append("p.payment_method = ?")
            params.append(payment_method)
    if created_by:
        filters.append("p.created_by = ?")
        params.append(created_by)

    rows = query_all(
        f"""
        SELECT
            p.id AS payment_id,
            p.receivable_id,
            p.amount AS received_amount,
            p.payment_date,
            COALESCE(NULLIF(TRIM(p.payment_method), ''), 'Não informado') AS payment_method,
            COALESCE(p.notes, '') AS payment_notes,
            COALESCE(NULLIF(TRIM(p.created_by), ''), 'Não informado') AS created_by,
            p.created_at AS payment_created_at,
            r.description,
            r.amount AS installment_amount,
            r.due_date,
            r.installment_number,
            r.installment_total,
            r.receivable_group_id,
            c.id AS client_id,
            c.name AS client_name
        FROM receivable_payments p
        JOIN receivables r ON r.id = p.receivable_id
        JOIN clients c ON c.id = r.client_id
        WHERE {' AND '.join(filters)}
        ORDER BY p.payment_date DESC, p.id DESC
        """,
        params,
    )

    result = []
    client_totals = {}
    method_totals = {}
    day_totals = {}

    for row in rows:
        received_amount = float(row["received_amount"] or 0)
        item = {
            "payment_id": row["payment_id"],
            "receivable_id": row["receivable_id"],
            "received_amount": received_amount,
            "payment_date": row["payment_date"],
            "payment_method": row["payment_method"],
            "payment_notes": row["payment_notes"],
            "created_by": row["created_by"],
            "payment_created_at": row["payment_created_at"],
            "description": row["description"],
            "installment_amount": float(row["installment_amount"] or 0),
            "due_date": row["due_date"],
            "installment_number": int(row["installment_number"] or 1),
            "installment_total": int(row["installment_total"] or 1),
            "receivable_group_id": row["receivable_group_id"],
            "client_id": row["client_id"],
            "client_name": row["client_name"],
        }
        result.append(item)

        client_data = client_totals.setdefault(
            item["client_id"],
            {"client_id": item["client_id"], "client_name": item["client_name"], "amount": 0.0, "count": 0},
        )
        client_data["amount"] += received_amount
        client_data["count"] += 1

        method_data = method_totals.setdefault(
            item["payment_method"],
            {"payment_method": item["payment_method"], "amount": 0.0, "count": 0},
        )
        method_data["amount"] += received_amount
        method_data["count"] += 1

        day_data = day_totals.setdefault(
            item["payment_date"],
            {"payment_date": item["payment_date"], "amount": 0.0, "count": 0},
        )
        day_data["amount"] += received_amount
        day_data["count"] += 1

    total_received = sum(item["received_amount"] for item in result)
    receipts_count = len(result)
    unique_clients = len(client_totals)
    average_receipt = total_received / receipts_count if receipts_count else 0.0

    return {
        "rows": result,
        "summary": {
            "total_received": total_received,
            "receipts_count": receipts_count,
            "unique_clients": unique_clients,
            "average_receipt": average_receipt,
        },
        "by_client": sorted(client_totals.values(), key=lambda item: (-item["amount"], item["client_name"].casefold())),
        "by_method": sorted(method_totals.values(), key=lambda item: (-item["amount"], item["payment_method"].casefold())),
        "by_day": sorted(day_totals.values(), key=lambda item: item["payment_date"]),
    }


def human_file_size(size_bytes):
    try:
        size = float(size_bytes or 0)
    except (TypeError, ValueError):
        size = 0
    units = ["B", "KB", "MB", "GB"]
    unit_index = 0
    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1
    if unit_index == 0:
        return f"{int(size)} {units[unit_index]}"
    return f"{size:.1f} {units[unit_index]}".replace(".", ",")


def is_safe_backup_filename(filename):
    filename = filename or ""
    return (
        filename == secure_filename(filename)
        and (filename.startswith("backup_banco_gestao360_") or filename.startswith("backup_completo_gestao360_"))
        and (filename.endswith(".db") or filename.endswith(".zip"))
    )


def list_backup_files():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    backups = []
    for filename in os.listdir(BACKUP_DIR):
        if not is_safe_backup_filename(filename):
            continue
        path = os.path.join(BACKUP_DIR, filename)
        if not os.path.isfile(path):
            continue
        stat = os.stat(path)
        backups.append({
            "filename": filename,
            "kind": "Completo" if filename.endswith(".zip") else "Banco de dados",
            "size": stat.st_size,
            "size_label": human_file_size(stat.st_size),
            "created_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        })
    backups.sort(key=lambda item: item["created_at"], reverse=True)
    return backups


def create_database_backup():
    if using_remote_database():
        raise RuntimeError("O banco em nuvem deve ser protegido pelos backups do provedor Turso.")
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"backup_banco_gestao360_{stamp}.db"
    destination = os.path.join(BACKUP_DIR, filename)

    # Usa a API de backup do SQLite para gerar uma cópia consistente mesmo com o sistema aberto.
    source = sqlite3.connect(DB_PATH)
    target = sqlite3.connect(destination)
    try:
        source.backup(target)
    finally:
        target.close()
        source.close()
    return filename


def create_full_backup():
    if using_remote_database():
        raise RuntimeError("O backup local completo não está disponível com banco em nuvem.")
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"backup_completo_gestao360_{stamp}.zip"
    destination = os.path.join(BACKUP_DIR, filename)
    temp_db_path = os.path.join(BACKUP_DIR, f"_gestao360_temp_{stamp}.db")

    source = sqlite3.connect(DB_PATH)
    target = sqlite3.connect(temp_db_path)
    try:
        source.backup(target)
    finally:
        target.close()
        source.close()

    try:
        with zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED) as zf:
            if os.path.exists(temp_db_path):
                zf.write(temp_db_path, arcname="gestao360.db")
            if os.path.isdir(os.path.join(BASE_DIR, "uploads")):
                uploads_root = os.path.join(BASE_DIR, "uploads")
                for root, _dirs, files in os.walk(uploads_root):
                    for file in files:
                        path = os.path.join(root, file)
                        arcname = os.path.relpath(path, BASE_DIR)
                        zf.write(path, arcname=arcname)
            zf.writestr(
                "LEIA-ME.txt",
                "Backup completo do Gestão360 Contábil.\n"
                "Contém o banco gestao360.db e anexos enviados nas atividades.\n"
                "Guarde este arquivo em local seguro.\n",
            )
    finally:
        try:
            os.remove(temp_db_path)
        except OSError:
            pass
    return filename


@app.context_processor
def inject_globals():
    # A página de configuração da Vercel precisa funcionar antes de o banco existir.
    if IS_VERCEL and globals().get("MISSING_VERCEL_ENV"):
        return {
            "app_name": APP_NAME,
            "money": money,
            "hours": hours,
            "month_label": month_label,
            "can_view_financial": False,
            "current_month": current_month(),
        }
    return {
        "app_name": APP_NAME,
        "money": money,
        "hours": hours,
        "month_label": month_label,
        "can_view_financial": can_view_financial(),
        "current_month": current_month(),
    }


@app.route("/primeiro-acesso", methods=["GET", "POST"])
def first_access():
    if system_has_users():
        return redirect(url_for("login"))

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        confirmation = request.form.get("password_confirmation") or ""
        registration_code = (request.form.get("registration_code") or "").strip()

        error = validate_password(password)
        if not name or not email:
            flash("Informe seu nome e e-mail.", "error")
        elif error:
            flash(error, "error")
        elif password != confirmation:
            flash("A confirmação da senha não confere.", "error")
        elif len(registration_code) < 6:
            flash("Crie um código do escritório com pelo menos 6 caracteres.", "error")
        else:
            try:
                with get_db() as conn:
                    cur = conn.execute(
                        "INSERT INTO users (name, email, password_hash, role) VALUES (?, ?, ?, ?)",
                        (name, email, generate_password_hash(password), "Administrador"),
                    )
                    user_id = cur.lastrowid
                    conn.execute(
                        "INSERT OR REPLACE INTO settings (key, value) VALUES ('registration_code_hash', ?)",
                        (generate_password_hash(registration_code),),
                    )
                    conn.execute(
                        "INSERT OR REPLACE INTO settings (key, value) VALUES ('allow_self_registration', '1')"
                    )
                    conn.commit()
                user = query_one("SELECT * FROM users WHERE id = ?", (user_id,))
                login_user_session(user)
                audit_log("Configurou primeiro acesso", "Usuário", user_id, name, "Primeiro administrador criado")
                flash("Conta administradora criada com sucesso.", "success")
                return redirect(url_for("dashboard"))
            except (sqlite3.IntegrityError, ValueError):
                flash("Já existe uma conta com esse e-mail.", "error")

    return render_template("first_access.html")


@app.route("/criar-conta", methods=["GET", "POST"])
def register():
    if not system_has_users():
        return redirect(url_for("first_access"))

    code_hash = setting_value("registration_code_hash", "")
    registration_enabled = setting_value("allow_self_registration", "1") == "1" and bool(code_hash)
    if not registration_enabled:
        flash("A criação de novas contas está desativada. Procure o administrador.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        confirmation = request.form.get("password_confirmation") or ""
        registration_code = request.form.get("registration_code") or ""
        error = validate_password(password)

        if not name or not email:
            flash("Informe seu nome e e-mail.", "error")
        elif error:
            flash(error, "error")
        elif password != confirmation:
            flash("A confirmação da senha não confere.", "error")
        elif not code_hash or not check_password_hash(code_hash, registration_code):
            flash("Código do escritório inválido.", "error")
        else:
            try:
                new_user_id = execute(
                    "INSERT INTO users (name, email, password_hash, role) VALUES (?, ?, ?, ?)",
                    (name, email, generate_password_hash(password), "Colaborador"),
                )
                user = query_one("SELECT * FROM users WHERE id = ?", (new_user_id,))
                login_user_session(user)
                audit_log("Criou a própria conta", "Usuário", new_user_id, name, f"E-mail: {email}; perfil inicial: Colaborador")
                flash("Sua conta foi criada com sucesso.", "success")
                return redirect(url_for("dashboard"))
            except (sqlite3.IntegrityError, ValueError):
                flash("Já existe uma conta com esse e-mail.", "error")

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if not system_has_users():
        return redirect(url_for("first_access"))
    if session.get("user_id"):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = query_one("SELECT * FROM users WHERE email = ? AND active = 1", (email,))
        if user and check_password_hash(user["password_hash"], password):
            login_user_session(user)
            return redirect(url_for("dashboard"))
        flash("E-mail ou senha inválidos.", "error")
    return render_template(
        "login.html",
        registration_enabled=(
            setting_value("allow_self_registration", "1") == "1"
            and bool(setting_value("registration_code_hash", ""))
        ),
    )


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    month = request.args.get("month") or current_month()
    start, end = month_range(month)

    active_clients = query_one("SELECT COUNT(*) AS total FROM clients WHERE status = 'Ativo'")["total"]
    hours_month = query_one(
        "SELECT COALESCE(SUM(time_spent), 0) AS total FROM activities WHERE activity_date >= ? AND activity_date < ?",
        (start, end),
    )["total"]

    report = build_monthly_report(month)
    client_most_hours = max(report, key=lambda x: x["total_hours"], default=None)
    profitable_clients = [r for r in report if r["value_per_hour"] is not None]
    most_profitable = max(profitable_clients, key=lambda x: x["value_per_hour"], default=None)
    risk_client = min(profitable_clients, key=lambda x: x["value_per_hour"], default=None)

    total_to_receive = total_received = total_overdue = 0
    status_counts = {s: 0 for s in RECEIVABLE_STATUS}
    if can_view_financial():
        total_received = query_one(
            "SELECT COALESCE(SUM(amount), 0) AS total FROM receivable_payments WHERE payment_date >= ? AND payment_date < ?",
            (start, end),
        )["total"]
        receivables = query_all(
            """
            SELECT *
            FROM receivables r
            WHERE EXISTS (
                    SELECT 1 FROM receivable_payments p
                    WHERE p.receivable_id = r.id
                      AND p.payment_date >= ?
                      AND p.payment_date < ?
                  )
               OR (r.status != 'Cancelado' AND (r.amount - COALESCE(r.paid_amount, 0)) > 0 AND r.due_date < ?)
               OR (r.status = 'Cancelado' AND r.due_date >= ? AND r.due_date < ?)
            """,
            (start, end, end, start, end),
        )
        for r in receivables:
            st = computed_receivable_status(r)
            status_counts[st] = status_counts.get(st, 0) + 1
            remaining = receivable_remaining_amount(r)
            if st != "Cancelado" and r["due_date"] and r["due_date"] < end and remaining > 0:
                total_to_receive += remaining
            if is_receivable_overdue(r):
                total_overdue += remaining

    chart_hours = [r for r in sorted(report, key=lambda x: x["total_hours"], reverse=True)[:8] if r["total_hours"] > 0]
    ranking_received = sorted(report, key=lambda x: x["received_amount"], reverse=True)[:8]

    today = date.today().isoformat()
    open_activity_filter = "COALESCE(a.status, 'Pendente') != 'Concluído'"

    task_total_month = query_one(
        "SELECT COUNT(*) AS total FROM activities a WHERE a.activity_date >= ? AND a.activity_date < ?",
        (start, end),
    )["total"]
    task_open_month = query_one(
        f"SELECT COUNT(*) AS total FROM activities a WHERE a.activity_date >= ? AND a.activity_date < ? AND {open_activity_filter}",
        (start, end),
    )["total"]
    task_completed_month = query_one(
        "SELECT COUNT(*) AS total FROM activities a WHERE a.activity_date >= ? AND a.activity_date < ? AND a.status = 'Concluído'",
        (start, end),
    )["total"]
    tasks_overdue_count = query_one(
        f"SELECT COUNT(*) AS total FROM activities a WHERE {open_activity_filter} AND a.due_date IS NOT NULL AND a.due_date < ?",
        (today,),
    )["total"]
    tasks_due_today_count = query_one(
        f"SELECT COUNT(*) AS total FROM activities a WHERE {open_activity_filter} AND a.due_date = ?",
        (today,),
    )["total"]
    tasks_waiting_count = query_one(
        f"""
        SELECT COUNT(*) AS total
        FROM activities a
        WHERE {open_activity_filter}
          AND a.status IN ('Aguardando cliente', 'Aguardando documentos')
        """,
    )["total"]
    high_priority_open_count = query_one(
        f"""
        SELECT COUNT(*) AS total
        FROM activities a
        WHERE {open_activity_filter}
          AND COALESCE(a.priority, 'Média') IN ('Urgente', 'Alta')
        """,
    )["total"]

    task_status_counts = {s: 0 for s in ACTIVITY_STATUS}
    for row in query_all(
        """
        SELECT COALESCE(status, 'Pendente') AS status, COUNT(*) AS total
        FROM activities
        WHERE activity_date >= ? AND activity_date < ?
        GROUP BY COALESCE(status, 'Pendente')
        """,
        (start, end),
    ):
        task_status_counts[row["status"]] = int(row["total"] or 0)

    task_priority_counts = {p: 0 for p in ACTIVITY_PRIORITIES}
    for row in query_all(
        """
        SELECT COALESCE(priority, 'Média') AS priority, COUNT(*) AS total
        FROM activities
        WHERE activity_date >= ? AND activity_date < ?
        GROUP BY COALESCE(priority, 'Média')
        """,
        (start, end),
    ):
        task_priority_counts[row["priority"]] = int(row["total"] or 0)

    priority_case = """
        CASE COALESCE(a.priority, 'Média')
            WHEN 'Urgente' THEN 1
            WHEN 'Alta' THEN 2
            WHEN 'Média' THEN 3
            WHEN 'Baixa' THEN 4
            ELSE 5
        END
    """
    overdue_activities = query_all(
        f"""
        SELECT a.*, c.name AS client_name
        FROM activities a
        JOIN clients c ON c.id = a.client_id
        WHERE {open_activity_filter}
          AND a.due_date IS NOT NULL
          AND a.due_date < ?
        ORDER BY a.due_date ASC, {priority_case}, a.id DESC
        LIMIT 8
        """,
        (today,),
    )
    today_activities = query_all(
        f"""
        SELECT a.*, c.name AS client_name
        FROM activities a
        JOIN clients c ON c.id = a.client_id
        WHERE {open_activity_filter}
          AND a.due_date = ?
        ORDER BY {priority_case}, a.activity_date DESC, a.id DESC
        LIMIT 8
        """,
        (today,),
    )
    waiting_activities = query_all(
        f"""
        SELECT a.*, c.name AS client_name
        FROM activities a
        JOIN clients c ON c.id = a.client_id
        WHERE {open_activity_filter}
          AND a.status IN ('Aguardando cliente', 'Aguardando documentos')
        ORDER BY COALESCE(a.due_date, '9999-12-31') ASC, {priority_case}, a.id DESC
        LIMIT 8
        """
    )
    collaborator_ranking = query_all(
        f"""
        SELECT COALESCE(NULLIF(TRIM(a.collaborator), ''), 'Sem responsável') AS collaborator_name,
               COUNT(*) AS total_tasks,
               SUM(CASE WHEN {open_activity_filter} THEN 1 ELSE 0 END) AS open_tasks,
               SUM(CASE WHEN a.status = 'Concluído' THEN 1 ELSE 0 END) AS completed_tasks,
               SUM(CASE WHEN {open_activity_filter} AND a.due_date IS NOT NULL AND a.due_date < ? THEN 1 ELSE 0 END) AS overdue_tasks,
               COALESCE(SUM(a.time_spent), 0) AS total_hours
        FROM activities a
        WHERE a.activity_date >= ? AND a.activity_date < ?
        GROUP BY COALESCE(NULLIF(TRIM(a.collaborator), ''), 'Sem responsável')
        ORDER BY overdue_tasks DESC, open_tasks DESC, total_tasks DESC, total_hours DESC
        LIMIT 8
        """,
        (today, start, end),
    )

    return render_template(
        "dashboard.html",
        month=month,
        active_clients=active_clients,
        total_to_receive=total_to_receive,
        total_received=total_received,
        total_overdue=total_overdue,
        hours_month=hours_month,
        client_most_hours=client_most_hours,
        most_profitable=most_profitable,
        risk_client=risk_client,
        chart_hours=chart_hours,
        status_counts=status_counts,
        ranking_hours=chart_hours,
        ranking_received=ranking_received,
        task_total_month=task_total_month,
        task_open_month=task_open_month,
        task_completed_month=task_completed_month,
        tasks_overdue_count=tasks_overdue_count,
        tasks_due_today_count=tasks_due_today_count,
        tasks_waiting_count=tasks_waiting_count,
        high_priority_open_count=high_priority_open_count,
        task_status_counts=task_status_counts,
        task_priority_counts=task_priority_counts,
        overdue_activities=overdue_activities,
        today_activities=today_activities,
        waiting_activities=waiting_activities,
        collaborator_ranking=collaborator_ranking,
        today=today,
    )


@app.route("/clientes")
@login_required
def clients():
    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()
    filters = ["1=1"]
    params = []
    if search:
        filters.append("(name LIKE ? OR document LIKE ? OR responsible LIKE ?)")
        like = f"%{search}%"
        params += [like, like, like]
    if status:
        filters.append("status = ?")
        params.append(status)
    rows = query_all(f"SELECT * FROM clients WHERE {' AND '.join(filters)} ORDER BY name ASC", params)
    return render_template("clients.html", clients=rows, search=search, status=status, statuses=CLIENT_STATUS)


@app.route("/clientes/novo", methods=["GET", "POST"])
@login_required
@admin_required
def client_new():
    if request.method == "POST":
        client_name = request.form.get("name")
        new_client_id = execute(
            """
            INSERT INTO clients (name, document, client_type, phone, email, responsible, monthly_fee, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                client_name, request.form.get("document"), request.form.get("client_type"),
                request.form.get("phone"), request.form.get("email"), request.form.get("responsible"),
                parse_money(request.form.get("monthly_fee")), request.form.get("status"), request.form.get("notes"),
            ),
        )
        audit_log("Criou cliente", "Cliente", new_client_id, client_name, f"Status: {request.form.get('status') or '-'}")
        flash("Cliente cadastrado com sucesso.", "success")
        return redirect(url_for("clients"))
    return render_template("client_form.html", client=None, client_types=CLIENT_TYPES, statuses=CLIENT_STATUS)


@app.route("/clientes/<int:client_id>/editar", methods=["GET", "POST"])
@login_required
@admin_required
def client_edit(client_id):
    client = query_one("SELECT * FROM clients WHERE id = ?", (client_id,))
    if not client:
        flash("Cliente não encontrado.", "error")
        return redirect(url_for("clients"))
    if request.method == "POST":
        new_name = request.form.get("name")
        execute(
            """
            UPDATE clients
               SET name=?, document=?, client_type=?, phone=?, email=?, responsible=?, monthly_fee=?, status=?, notes=?
             WHERE id=?
            """,
            (
                new_name, request.form.get("document"), request.form.get("client_type"),
                request.form.get("phone"), request.form.get("email"), request.form.get("responsible"),
                parse_money(request.form.get("monthly_fee")), request.form.get("status"), request.form.get("notes"), client_id,
            ),
        )
        audit_log("Editou cliente", "Cliente", client_id, new_name, f"Nome anterior: {client['name']}")
        flash("Cliente atualizado com sucesso.", "success")
        return redirect(url_for("client_detail", client_id=client_id))
    return render_template("client_form.html", client=client, client_types=CLIENT_TYPES, statuses=CLIENT_STATUS)


@app.route("/clientes/<int:client_id>/excluir", methods=["POST"])
@login_required
@admin_required
def client_delete(client_id):
    client = query_one("SELECT * FROM clients WHERE id = ?", (client_id,))
    if not client:
        flash("Cliente não encontrado.", "error")
        return redirect(url_for("clients"))

    attachments = query_all(
        """
        SELECT af.stored_filename
        FROM activity_files af
        JOIN activities a ON a.id = af.activity_id
        WHERE a.client_id = ?
        """,
        (client_id,),
    )
    for item in attachments:
        try:
            os.remove(os.path.join(UPLOAD_DIR, item["stored_filename"]))
        except OSError:
            pass

    with get_db() as conn:
        # Apaga primeiro os dados vinculados para funcionar mesmo em bancos antigos.
        conn.execute("DELETE FROM activity_files WHERE activity_id IN (SELECT id FROM activities WHERE client_id = ?)", (client_id,))
        conn.execute("DELETE FROM activity_updates WHERE activity_id IN (SELECT id FROM activities WHERE client_id = ?)", (client_id,))
        conn.execute("DELETE FROM activity_time_entries WHERE activity_id IN (SELECT id FROM activities WHERE client_id = ?)", (client_id,))
        conn.execute("DELETE FROM activity_checklist_items WHERE activity_id IN (SELECT id FROM activities WHERE client_id = ?)", (client_id,))
        conn.execute("DELETE FROM activities WHERE client_id = ?", (client_id,))
        conn.execute("DELETE FROM recurring_activity_templates WHERE client_id = ?", (client_id,))
        conn.execute("DELETE FROM receivables WHERE client_id = ?", (client_id,))
        conn.execute("DELETE FROM clients WHERE id = ?", (client_id,))
        conn.commit()

    audit_log("Excluiu cliente", "Cliente", client_id, client["name"], "Cliente e dados vinculados removidos.")
    flash(f"Cliente {client['name']} excluído com sucesso.", "success")
    return redirect(url_for("clients"))


@app.route("/clientes/<int:client_id>")
@login_required
def client_detail(client_id):
    client = query_one("SELECT * FROM clients WHERE id = ?", (client_id,))
    if not client:
        flash("Cliente não encontrado.", "error")
        return redirect(url_for("clients"))
    month = request.args.get("month") or current_month()
    start, end = month_range(month)
    activities = query_all(
        "SELECT * FROM activities WHERE client_id = ? ORDER BY activity_date DESC, id DESC LIMIT 200", (client_id,)
    )
    monthly_hours = query_one(
        "SELECT COALESCE(SUM(time_spent), 0) AS total FROM activities WHERE client_id = ? AND activity_date >= ? AND activity_date < ?",
        (client_id, start, end),
    )["total"]
    activity_count = query_one(
        "SELECT COUNT(*) AS total FROM activities WHERE client_id = ? AND activity_date >= ? AND activity_date < ?",
        (client_id, start, end),
    )["total"]
    receivables = []
    received = overdue = to_receive = 0
    if can_view_financial():
        receivables = query_all("SELECT * FROM receivables WHERE client_id = ? ORDER BY due_date DESC, id DESC LIMIT 100", (client_id,))
        received = query_one(
            """
            SELECT COALESCE(SUM(p.amount), 0) AS total
            FROM receivable_payments p
            JOIN receivables r ON r.id = p.receivable_id
            WHERE r.client_id = ? AND p.payment_date >= ? AND p.payment_date < ?
            """,
            (client_id, start, end),
        )["total"]
        for r in receivables:
            remaining = receivable_remaining_amount(r)
            if is_receivable_overdue(r):
                overdue += remaining
            if computed_receivable_status(r) != "Cancelado" and remaining > 0:
                to_receive += remaining
    value_per_hour = None if float(monthly_hours or 0) <= 0 else float(client["monthly_fee"] or 0) / float(monthly_hours or 0)
    return render_template(
        "client_detail.html", client=client, month=month, activities=activities,
        monthly_hours=monthly_hours, activity_count=activity_count, receivables=receivables,
        received=received, overdue=overdue, to_receive=to_receive, value_per_hour=value_per_hour,
        rentability=rentability_label(value_per_hour), service_types=SERVICE_TYPES, activity_statuses=ACTIVITY_STATUS,
        receivable_statuses=RECEIVABLE_STATUS, payment_methods=PAYMENT_METHODS,
        today=date.today().isoformat(), computed_receivable_status=computed_receivable_status,
        receivable_paid_amount=receivable_paid_amount, receivable_remaining_amount=receivable_remaining_amount,
        can_receive_payment=can_receive_payment, has_receivable_payment=has_receivable_payment
    )


@app.route("/modelos-checklist", methods=["GET", "POST"])
@login_required
@admin_required
def checklist_templates():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        service_type = request.form.get("service_type", "").strip() or None
        description = request.form.get("description", "").strip()
        initial_items = parse_checklist_lines(request.form.get("items_text", ""))

        if not name:
            flash("Informe um nome para o modelo de checklist.", "error")
            return redirect(url_for("checklist_templates"))

        try:
            with get_db() as conn:
                cur = conn.execute(
                    """
                    INSERT INTO checklist_templates (name, description, service_type, created_by)
                    VALUES (?, ?, ?, ?)
                    """,
                    (name, description, service_type, session.get("name")),
                )
                template_id = cur.lastrowid
                for index, title in enumerate(initial_items, start=1):
                    conn.execute(
                        """
                        INSERT INTO checklist_template_items (template_id, title, sort_order)
                        VALUES (?, ?, ?)
                        """,
                        (template_id, title, index),
                    )
                add_audit_log_conn(conn, "Criou modelo de checklist", "Modelo de checklist", template_id, name, f"Itens iniciais: {len(initial_items)}")
                conn.commit()
            flash("Modelo de checklist criado com sucesso.", "success")
        except (sqlite3.IntegrityError, ValueError):
            flash("Já existe um modelo com esse nome. Use outro nome.", "error")
        return redirect(url_for("checklist_templates"))

    templates = query_all(
        """
        SELECT t.*,
               COUNT(i.id) AS items_count
        FROM checklist_templates t
        LEFT JOIN checklist_template_items i ON i.template_id = t.id
        GROUP BY t.id
        ORDER BY t.active DESC, t.name COLLATE NOCASE
        """
    )
    template_items = {}
    if templates:
        ids = [int(t["id"]) for t in templates]
        placeholders = ",".join("?" for _ in ids)
        rows = query_all(
            f"""
            SELECT *
            FROM checklist_template_items
            WHERE template_id IN ({placeholders})
            ORDER BY template_id ASC, sort_order ASC, id ASC
            """,
            ids,
        )
        for row in rows:
            template_items.setdefault(int(row["template_id"]), []).append(row)

    return render_template(
        "checklist_templates.html",
        templates=templates,
        template_items=template_items,
        service_types=SERVICE_TYPES,
    )


@app.route("/modelos-checklist/<int:template_id>/atualizar", methods=["POST"])
@login_required
@admin_required
def checklist_template_update(template_id):
    template = query_one("SELECT * FROM checklist_templates WHERE id = ?", (template_id,))
    if not template:
        flash("Modelo de checklist não encontrado.", "error")
        return redirect(url_for("checklist_templates"))

    name = request.form.get("name", "").strip()
    if not name:
        flash("O nome do modelo não pode ficar vazio.", "error")
        return redirect(url_for("checklist_templates"))
    service_type = request.form.get("service_type", "").strip() or None
    active = 1 if request.form.get("active") == "1" else 0
    try:
        execute(
            """
            UPDATE checklist_templates
               SET name = ?, description = ?, service_type = ?, active = ?, updated_at = CURRENT_TIMESTAMP
             WHERE id = ?
            """,
            (name, request.form.get("description", "").strip(), service_type, active, template_id),
        )
        audit_log("Editou modelo de checklist", "Modelo de checklist", template_id, name, f"Ativo: {'sim' if active else 'não'}")
        flash("Modelo atualizado.", "success")
    except (sqlite3.IntegrityError, ValueError):
        flash("Já existe outro modelo com esse nome.", "error")
    return redirect(url_for("checklist_templates") + f"#modelo-{template_id}")


@app.route("/modelos-checklist/<int:template_id>/item/adicionar", methods=["POST"])
@login_required
@admin_required
def checklist_template_item_add(template_id):
    template = query_one("SELECT * FROM checklist_templates WHERE id = ?", (template_id,))
    if not template:
        flash("Modelo de checklist não encontrado.", "error")
        return redirect(url_for("checklist_templates"))
    title = request.form.get("title", "").strip()
    if not title:
        flash("Escreva o item antes de adicionar.", "error")
        return redirect(url_for("checklist_templates") + f"#modelo-{template_id}")
    max_order = query_one(
        "SELECT COALESCE(MAX(sort_order), 0) AS max_order FROM checklist_template_items WHERE template_id = ?",
        (template_id,),
    )["max_order"]
    execute(
        """
        INSERT INTO checklist_template_items (template_id, title, sort_order)
        VALUES (?, ?, ?)
        """,
        (template_id, title, int(max_order or 0) + 1),
    )
    audit_log("Adicionou item em modelo", "Modelo de checklist", template_id, template["name"], title)
    flash("Item adicionado ao modelo.", "success")
    return redirect(url_for("checklist_templates") + f"#modelo-{template_id}")


@app.route("/modelos-checklist/item/<int:item_id>/editar", methods=["POST"])
@login_required
@admin_required
def checklist_template_item_edit(item_id):
    item = query_one("SELECT * FROM checklist_template_items WHERE id = ?", (item_id,))
    if not item:
        flash("Item do modelo não encontrado.", "error")
        return redirect(url_for("checklist_templates"))
    title = request.form.get("title", "").strip()
    if not title:
        flash("O item do modelo não pode ficar vazio.", "error")
        return redirect(url_for("checklist_templates") + f"#modelo-{item['template_id']}")
    execute("UPDATE checklist_template_items SET title = ? WHERE id = ?", (title, item_id))
    audit_log("Editou item de modelo", "Modelo de checklist", item["template_id"], title, f"Anterior: {item['title']}")
    flash("Item do modelo atualizado.", "success")
    return redirect(url_for("checklist_templates") + f"#modelo-{item['template_id']}")


@app.route("/modelos-checklist/item/<int:item_id>/excluir", methods=["POST"])
@login_required
@admin_required
def checklist_template_item_delete(item_id):
    item = query_one("SELECT * FROM checklist_template_items WHERE id = ?", (item_id,))
    if not item:
        flash("Item do modelo não encontrado.", "error")
        return redirect(url_for("checklist_templates"))
    template_id = item["template_id"]
    execute("DELETE FROM checklist_template_items WHERE id = ?", (item_id,))
    audit_log("Excluiu item de modelo", "Modelo de checklist", template_id, item["title"], "Item removido do modelo")
    flash("Item removido do modelo.", "success")
    return redirect(url_for("checklist_templates") + f"#modelo-{template_id}")


@app.route("/modelos-checklist/<int:template_id>/excluir", methods=["POST"])
@login_required
@admin_required
def checklist_template_delete(template_id):
    template = query_one("SELECT * FROM checklist_templates WHERE id = ?", (template_id,))
    if not template:
        flash("Modelo de checklist não encontrado.", "error")
        return redirect(url_for("checklist_templates"))
    execute("DELETE FROM checklist_templates WHERE id = ?", (template_id,))
    audit_log("Excluiu modelo de checklist", "Modelo de checklist", template_id, template["name"], "Modelo e itens removidos")
    flash("Modelo de checklist excluído.", "success")
    return redirect(url_for("checklist_templates"))


@app.route("/rotinas-mensais", methods=["GET", "POST"])
@login_required
@admin_required
def recurring_tasks():
    if request.method == "POST":
        client_id_raw = request.form.get("client_id", "").strip()
        client_id = int(client_id_raw) if client_id_raw else None
        service_type = request.form.get("service_type", "").strip()
        description = request.form.get("description", "").strip()
        collaborator = request.form.get("collaborator", "").strip() or session.get("name")
        priority = request.form.get("priority", "Média")
        if priority not in ACTIVITY_PRIORITIES:
            priority = "Média"
        try:
            activity_day = int(request.form.get("activity_day") or 1)
        except ValueError:
            activity_day = 1
        try:
            due_day = int(request.form.get("due_day") or 0)
        except ValueError:
            due_day = 0
        checklist_template_id_raw = request.form.get("checklist_template_id", "").strip()
        checklist_template_id = int(checklist_template_id_raw) if checklist_template_id_raw else None

        if not service_type or not description:
            flash("Informe o tipo de serviço e a descrição da rotina.", "error")
            return redirect(url_for("recurring_tasks"))

        activity_day = min(max(activity_day, 1), 31)
        due_day = min(max(due_day, 1), 31) if due_day else None
        execute(
            """
            INSERT INTO recurring_activity_templates
                (client_id, service_type, description, collaborator, priority, activity_day, due_day, checklist_template_id, notes, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                client_id,
                service_type,
                description,
                collaborator,
                priority,
                activity_day,
                due_day,
                checklist_template_id,
                request.form.get("notes", "").strip(),
                session.get("name"),
            ),
        )
        audit_log("Criou rotina mensal", "Rotina mensal", None, description, f"Serviço: {service_type}; prioridade: {priority}")
        flash("Rotina mensal cadastrada com sucesso.", "success")
        return redirect(url_for("recurring_tasks"))

    month = request.args.get("month") or current_month()
    clients_list = query_all("SELECT id, name FROM clients WHERE status != 'Inativo' ORDER BY name COLLATE NOCASE")
    checklist_templates_list = get_active_checklist_templates()
    templates = query_all(
        """
        SELECT rt.*, c.name AS client_name, ct.name AS checklist_template_name,
               COALESCE((
                   SELECT COUNT(*)
                   FROM activities a
                   WHERE a.recurrence_template_id = rt.id
                     AND a.recurrence_month = ?
               ), 0) AS generated_in_month
        FROM recurring_activity_templates rt
        LEFT JOIN clients c ON c.id = rt.client_id
        LEFT JOIN checklist_templates ct ON ct.id = rt.checklist_template_id
        ORDER BY rt.active DESC, COALESCE(c.name, 'Todos os clientes ativos') COLLATE NOCASE, rt.service_type, rt.description
        """,
        (month,),
    )
    active_count = sum(1 for t in templates if int(t["active"] or 0) == 1)
    generated_count = sum(int(t["generated_in_month"] or 0) for t in templates)
    return render_template(
        "recurring_tasks.html",
        month=month,
        clients=clients_list,
        checklist_templates=checklist_templates_list,
        templates=templates,
        service_types=SERVICE_TYPES,
        activity_priorities=ACTIVITY_PRIORITIES,
        active_count=active_count,
        generated_count=generated_count,
        today=date.today().isoformat(),
    )


@app.route("/rotinas-mensais/gerar", methods=["POST"])
@login_required
@admin_required
def recurring_tasks_generate():
    month = request.form.get("month") or current_month()
    result = generate_recurring_activities_for_month(month)
    audit_log("Gerou rotinas do mês", "Rotina mensal", None, month_label(month), f"Criadas: {result['created']}; ignoradas: {result['skipped']}")
    if result["created"]:
        flash(
            f"Rotinas de {month_label(month)} geradas: {result['created']} atividade(s) criada(s). "
            f"{result['skipped']} já existiam e foram ignoradas.",
            "success",
        )
    else:
        flash(
            f"Nenhuma atividade nova foi criada para {month_label(month)}. "
            f"{result['skipped']} rotina(s) já estavam geradas.",
            "error",
        )
    return redirect(url_for("recurring_tasks", month=month))


@app.route("/rotinas-mensais/<int:template_id>/atualizar", methods=["POST"])
@login_required
@admin_required
def recurring_task_update(template_id):
    template = query_one("SELECT * FROM recurring_activity_templates WHERE id = ?", (template_id,))
    if not template:
        flash("Rotina mensal não encontrada.", "error")
        return redirect(url_for("recurring_tasks"))

    client_id_raw = request.form.get("client_id", "").strip()
    client_id = int(client_id_raw) if client_id_raw else None
    priority = request.form.get("priority", "Média")
    if priority not in ACTIVITY_PRIORITIES:
        priority = "Média"
    try:
        activity_day = min(max(int(request.form.get("activity_day") or 1), 1), 31)
    except ValueError:
        activity_day = 1
    try:
        due_day_raw = int(request.form.get("due_day") or 0)
        due_day = min(max(due_day_raw, 1), 31) if due_day_raw else None
    except ValueError:
        due_day = None
    checklist_template_id_raw = request.form.get("checklist_template_id", "").strip()
    checklist_template_id = int(checklist_template_id_raw) if checklist_template_id_raw else None
    active = 1 if request.form.get("active") == "1" else 0
    description = request.form.get("description", "").strip()
    service_type = request.form.get("service_type", "").strip()
    if not description or not service_type:
        flash("Tipo de serviço e descrição são obrigatórios.", "error")
        return redirect(url_for("recurring_tasks") + f"#rotina-{template_id}")

    execute(
        """
        UPDATE recurring_activity_templates
           SET client_id = ?, service_type = ?, description = ?, collaborator = ?, priority = ?,
               activity_day = ?, due_day = ?, checklist_template_id = ?, active = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
         WHERE id = ?
        """,
        (
            client_id,
            service_type,
            description,
            request.form.get("collaborator", "").strip(),
            priority,
            activity_day,
            due_day,
            checklist_template_id,
            active,
            request.form.get("notes", "").strip(),
            template_id,
        ),
    )
    audit_log("Editou rotina mensal", "Rotina mensal", template_id, description, f"Ativa: {'sim' if active else 'não'}")
    flash("Rotina mensal atualizada.", "success")
    return redirect(url_for("recurring_tasks", month=request.form.get("month") or current_month()) + f"#rotina-{template_id}")


@app.route("/rotinas-mensais/<int:template_id>/excluir", methods=["POST"])
@login_required
@admin_required
def recurring_task_delete(template_id):
    template = query_one("SELECT * FROM recurring_activity_templates WHERE id = ?", (template_id,))
    if not template:
        flash("Rotina mensal não encontrada.", "error")
        return redirect(url_for("recurring_tasks"))
    execute("DELETE FROM recurring_activity_templates WHERE id = ?", (template_id,))
    audit_log("Excluiu rotina mensal", "Rotina mensal", template_id, template["description"], "Atividades já geradas foram mantidas")
    flash("Rotina mensal excluída. As atividades já geradas permanecem no histórico.", "success")
    return redirect(url_for("recurring_tasks"))


@app.route("/atividades", methods=["GET", "POST"])
@login_required
def activities():
    if request.method == "POST":
        initial_hours = float(request.form.get("time_spent") or 0)
        status = request.form.get("status") or "Pendente"
        if status not in ACTIVITY_STATUS:
            status = "Pendente"
        priority = request.form.get("priority") or "Média"
        if priority not in ACTIVITY_PRIORITIES:
            priority = "Média"
        new_activity_id = execute(
            """
            INSERT INTO activities (client_id, activity_date, service_type, description, collaborator, time_spent, status, due_date, priority, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(request.form.get("client_id")), request.form.get("activity_date"), request.form.get("service_type"),
                request.form.get("description"), request.form.get("collaborator"), initial_hours,
                status, request.form.get("due_date") or None, priority, request.form.get("notes"),
            ),
        )
        seconds = int(round(initial_hours * 3600))
        if seconds > 0:
            execute(
                """
                INSERT INTO activity_time_entries (activity_id, user_id, user_name, seconds)
                VALUES (?, ?, ?, ?)
                """,
                (new_activity_id, session.get("user_id"), session.get("name"), seconds),
            )
        audit_log("Criou atividade", "Atividade", new_activity_id, request.form.get("description"), f"Status: {status}; prioridade: {priority}")
        flash("Atividade registrada com sucesso. Agora você pode acompanhar a ficha da atividade.", "success")
        return redirect(url_for("activities"))

    month = request.args.get("month") or current_month()
    client_id = request.args.get("client_id", "")
    collaborator = request.args.get("collaborator", "").strip()
    service_type = request.args.get("service_type", "").strip()
    status_filter = request.args.get("status", "").strip()
    priority_filter = request.args.get("priority", "").strip()
    start, end = month_range(month)
    filters = ["a.activity_date >= ?", "a.activity_date < ?"]
    params = [start, end]
    if client_id:
        filters.append("a.client_id = ?")
        params.append(client_id)
    if collaborator:
        filters.append("a.collaborator LIKE ?")
        params.append(f"%{collaborator}%")
    if service_type:
        filters.append("a.service_type = ?")
        params.append(service_type)
    if status_filter:
        filters.append("a.status = ?")
        params.append(status_filter)
    if priority_filter:
        filters.append("COALESCE(a.priority, 'Média') = ?")
        params.append(priority_filter)
    rows = query_all(
        f"""
        SELECT a.*, c.name AS client_name
        FROM activities a
        JOIN clients c ON c.id = a.client_id
        WHERE {' AND '.join(filters)}
        ORDER BY
            CASE COALESCE(a.priority, 'Média')
                WHEN 'Urgente' THEN 1
                WHEN 'Alta' THEN 2
                WHEN 'Média' THEN 3
                WHEN 'Baixa' THEN 4
                ELSE 5
            END,
            COALESCE(a.due_date, '9999-12-31') ASC,
            a.activity_date DESC,
            a.id DESC
        """,
        params,
    )
    checklist_stats = get_activity_checklist_stats([int(r["id"]) for r in rows])
    clients_list = query_all("SELECT id, name FROM clients WHERE status != 'Inativo' ORDER BY name")
    total_hours = sum(float(r["time_spent"] or 0) for r in rows)
    kanban_counts = {s: 0 for s in ACTIVITY_STATUS}
    priority_counts = {p: 0 for p in ACTIVITY_PRIORITIES}
    overdue_count = 0
    for row in rows:
        kanban_counts[row["status"]] = kanban_counts.get(row["status"], 0) + 1
        priority_counts[row["priority"] or "Média"] = priority_counts.get(row["priority"] or "Média", 0) + 1
        if activity_is_overdue(row):
            overdue_count += 1
    return render_template(
        "activities.html", activities=rows, clients=clients_list, month=month,
        client_id=str(client_id), collaborator=collaborator, service_type=service_type,
        status_filter=status_filter, priority_filter=priority_filter,
        service_types=SERVICE_TYPES, activity_statuses=ACTIVITY_STATUS, activity_priorities=ACTIVITY_PRIORITIES,
        total_hours=total_hours, kanban_counts=kanban_counts, priority_counts=priority_counts,
        overdue_count=overdue_count, today=date.today().isoformat(), selected_client=request.args.get("selected_client"),
        activity_next_status=activity_next_status, activity_is_overdue=activity_is_overdue,
        checklist_stats=checklist_stats
    )


@app.route("/atividades/<int:activity_id>")
@login_required
def activity_detail(activity_id):
    activity = get_activity_or_404(activity_id)
    if not activity:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("activities"))

    time_entries = query_all(
        "SELECT * FROM activity_time_entries WHERE activity_id = ? ORDER BY created_at DESC, id DESC",
        (activity_id,),
    )
    updates = query_all(
        "SELECT * FROM activity_updates WHERE activity_id = ? ORDER BY created_at DESC, id DESC",
        (activity_id,),
    )
    files = query_all(
        "SELECT * FROM activity_files WHERE activity_id = ? ORDER BY created_at DESC, id DESC",
        (activity_id,),
    )
    checklist_items = query_all(
        """
        SELECT *
        FROM activity_checklist_items
        WHERE activity_id = ?
        ORDER BY is_done ASC, sort_order ASC, id ASC
        """,
        (activity_id,),
    )
    checklist_summary = checklist_progress(checklist_items)
    checklist_templates = get_active_checklist_templates()
    total_seconds = int(round(float(activity["time_spent"] or 0) * 3600))
    return render_template(
        "activity_detail.html",
        activity=activity,
        clients=query_all("SELECT id, name FROM clients WHERE status != 'Inativo' ORDER BY name"),
        service_types=SERVICE_TYPES,
        activity_statuses=ACTIVITY_STATUS,
        activity_priorities=ACTIVITY_PRIORITIES,
        activity_is_overdue=activity_is_overdue,
        time_entries=time_entries,
        updates=updates,
        files=files,
        checklist_items=checklist_items,
        checklist_summary=checklist_summary,
        checklist_templates=checklist_templates,
        total_seconds=total_seconds,
        format_seconds=format_seconds,
    )


@app.route("/atividades/<int:activity_id>/atualizar", methods=["POST"])
@login_required
def activity_update(activity_id):
    activity = get_activity_or_404(activity_id)
    if not activity:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("activities"))
    status = request.form.get("status") or activity["status"]
    if status not in ACTIVITY_STATUS:
        status = activity["status"]
    priority = request.form.get("priority") or "Média"
    if priority not in ACTIVITY_PRIORITIES:
        priority = "Média"
    execute(
        """
        UPDATE activities
           SET client_id=?, activity_date=?, service_type=?, description=?, collaborator=?, status=?, due_date=?, priority=?, notes=?, updated_at=CURRENT_TIMESTAMP
         WHERE id=?
        """,
        (
            int(request.form.get("client_id")),
            request.form.get("activity_date"),
            request.form.get("service_type"),
            request.form.get("description"),
            request.form.get("collaborator"),
            status,
            request.form.get("due_date") or None,
            priority,
            request.form.get("notes"),
            activity_id,
        ),
    )
    audit_log("Editou atividade", "Atividade", activity_id, request.form.get("description"), f"Status: {status}; prioridade: {priority}")
    flash("Ficha da atividade atualizada.", "success")
    return redirect(url_for("activity_detail", activity_id=activity_id))


@app.route("/atividades/<int:activity_id>/status", methods=["POST"])
@login_required
def activity_quick_status(activity_id):
    activity = get_activity_or_404(activity_id)
    if not activity:
        if request.accept_mimetypes.best == "application/json":
            return jsonify({"ok": False, "message": "Atividade não encontrada."}), 404
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("activities"))

    new_status = request.form.get("status", "").strip()
    if new_status not in ACTIVITY_STATUS:
        if request.accept_mimetypes.best == "application/json":
            return jsonify({"ok": False, "message": "Status inválido."}), 400
        flash("Status inválido.", "error")
        return redirect(request.referrer or url_for("activities"))

    old_status = activity["status"]
    if old_status != new_status:
        with get_db() as conn:
            conn.execute(
                "UPDATE activities SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (new_status, activity_id),
            )
            conn.execute(
                """
                INSERT INTO activity_updates (activity_id, user_id, user_name, message)
                VALUES (?, ?, ?, ?)
                """,
                (
                    activity_id,
                    session.get("user_id"),
                    session.get("name"),
                    f"Status alterado de {old_status} para {new_status} no quadro Kanban.",
                ),
            )
            add_audit_log_conn(conn, "Moveu atividade no Kanban", "Atividade", activity_id, activity["description"], f"{old_status} → {new_status}")
            conn.commit()

    if request.accept_mimetypes.best == "application/json":
        return jsonify({"ok": True, "status": new_status})
    flash(f"Atividade movida para {new_status}.", "success")
    return redirect(request.referrer or url_for("activities"))


@app.route("/atividades/<int:activity_id>/tempo", methods=["POST"])
@login_required
def activity_add_time(activity_id):
    activity = get_activity_or_404(activity_id)
    if not activity:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("activities"))
    try:
        seconds = int(float(request.form.get("elapsed_seconds") or 0))
    except ValueError:
        seconds = 0
    if seconds <= 0:
        flash("Nenhum tempo foi marcado no cronômetro.", "error")
        return redirect(url_for("activity_detail", activity_id=activity_id))

    added_hours = seconds_to_hours(seconds)
    with get_db() as conn:
        conn.execute(
            """
            INSERT INTO activity_time_entries (activity_id, user_id, user_name, seconds)
            VALUES (?, ?, ?, ?)
            """,
            (activity_id, session.get("user_id"), session.get("name"), seconds),
        )
        conn.execute(
            "UPDATE activities SET time_spent = COALESCE(time_spent, 0) + ?, status = CASE WHEN status = 'Pendente' THEN 'Em andamento' ELSE status END WHERE id = ?",
            (added_hours, activity_id),
        )
        add_audit_log_conn(conn, "Lançou tempo", "Atividade", activity_id, activity["description"], f"Tempo lançado: {format_seconds(seconds)}")
        conn.commit()
    flash(f"Tempo adicionado: {format_seconds(seconds)}.", "success")
    return redirect(url_for("activity_detail", activity_id=activity_id))


@app.route("/atividades/<int:activity_id>/andamento", methods=["POST"])
@login_required
def activity_add_update(activity_id):
    activity = get_activity_or_404(activity_id)
    if not activity:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("activities"))
    message = request.form.get("message", "").strip()
    if not message:
        flash("Escreva uma mensagem de andamento antes de salvar.", "error")
        return redirect(url_for("activity_detail", activity_id=activity_id))
    execute(
        """
        INSERT INTO activity_updates (activity_id, user_id, user_name, message)
        VALUES (?, ?, ?, ?)
        """,
        (activity_id, session.get("user_id"), session.get("name"), message),
    )
    audit_log("Registrou andamento", "Atividade", activity_id, activity["description"], message)
    flash("Andamento registrado.", "success")
    return redirect(url_for("activity_detail", activity_id=activity_id))


@app.route("/atividades/<int:activity_id>/checklist/aplicar-modelo", methods=["POST"])
@login_required
def activity_checklist_apply_template(activity_id):
    activity = get_activity_or_404(activity_id)
    if not activity:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("activities"))

    try:
        template_id = int(request.form.get("template_id") or 0)
    except ValueError:
        template_id = 0
    template = query_one("SELECT * FROM checklist_templates WHERE id = ? AND active = 1", (template_id,))
    if not template:
        flash("Selecione um modelo de checklist ativo.", "error")
        return redirect(url_for("activity_detail", activity_id=activity_id) + "#checklist")

    items = query_all(
        """
        SELECT *
        FROM checklist_template_items
        WHERE template_id = ?
        ORDER BY sort_order ASC, id ASC
        """,
        (template_id,),
    )
    if not items:
        flash("Esse modelo ainda não tem itens cadastrados.", "error")
        return redirect(url_for("activity_detail", activity_id=activity_id) + "#checklist")

    existing_rows = query_all("SELECT title FROM activity_checklist_items WHERE activity_id = ?", (activity_id,))
    existing_titles = {row["title"].strip().casefold() for row in existing_rows}
    max_order = query_one(
        "SELECT COALESCE(MAX(sort_order), 0) AS max_order FROM activity_checklist_items WHERE activity_id = ?",
        (activity_id,),
    )["max_order"]

    added = 0
    skipped = 0
    with get_db() as conn:
        sort_order = int(max_order or 0)
        for item in items:
            title = item["title"].strip()
            if not title:
                continue
            normalized = title.casefold()
            if normalized in existing_titles:
                skipped += 1
                continue
            sort_order += 1
            existing_titles.add(normalized)
            conn.execute(
                """
                INSERT INTO activity_checklist_items (activity_id, title, sort_order, created_by)
                VALUES (?, ?, ?, ?)
                """,
                (activity_id, title, sort_order, session.get("name")),
            )
            added += 1
        conn.execute(
            """
            INSERT INTO activity_updates (activity_id, user_id, user_name, message)
            VALUES (?, ?, ?, ?)
            """,
            (
                activity_id,
                session.get("user_id"),
                session.get("name"),
                f"Aplicou o modelo de checklist '{template['name']}' ({added} item(ns) adicionados).",
            ),
        )
        add_audit_log_conn(conn, "Aplicou modelo de checklist", "Atividade", activity_id, activity["description"], f"Modelo: {template['name']}; adicionados: {added}; pulados: {skipped}")
        conn.commit()

    if added:
        msg = f"Modelo aplicado: {added} item(ns) adicionado(s)."
        if skipped:
            msg += f" {skipped} item(ns) já existiam e não foram duplicados."
        flash(msg, "success")
    else:
        flash("Nenhum item novo foi adicionado, pois os itens do modelo já estavam no checklist.", "error")
    return redirect(url_for("activity_detail", activity_id=activity_id) + "#checklist")


@app.route("/atividades/<int:activity_id>/checklist/adicionar", methods=["POST"])
@login_required
def activity_checklist_add(activity_id):
    activity = get_activity_or_404(activity_id)
    if not activity:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("activities"))
    title = request.form.get("title", "").strip()
    if not title:
        flash("Escreva o item do checklist antes de adicionar.", "error")
        return redirect(url_for("activity_detail", activity_id=activity_id))
    max_order = query_one(
        "SELECT COALESCE(MAX(sort_order), 0) AS max_order FROM activity_checklist_items WHERE activity_id = ?",
        (activity_id,),
    )["max_order"]
    execute(
        """
        INSERT INTO activity_checklist_items (activity_id, title, sort_order, created_by)
        VALUES (?, ?, ?, ?)
        """,
        (activity_id, title, int(max_order or 0) + 1, session.get("name")),
    )
    add_activity_history(activity_id, f"Item adicionado ao checklist: {title}")
    audit_log("Adicionou item de checklist", "Atividade", activity_id, activity["description"], title)
    flash("Item adicionado ao checklist.", "success")
    return redirect(url_for("activity_detail", activity_id=activity_id) + "#checklist")


@app.route("/atividades/checklist/<int:item_id>/marcar", methods=["POST"])
@login_required
def activity_checklist_toggle(item_id):
    item = query_one("SELECT * FROM activity_checklist_items WHERE id = ?", (item_id,))
    if not item:
        flash("Item do checklist não encontrado.", "error")
        return redirect(url_for("activities"))
    done = 1 if request.form.get("is_done") == "1" else 0
    with get_db() as conn:
        conn.execute(
            """
            UPDATE activity_checklist_items
               SET is_done = ?,
                   completed_at = CASE WHEN ? = 1 THEN CURRENT_TIMESTAMP ELSE NULL END,
                   completed_by = CASE WHEN ? = 1 THEN ? ELSE NULL END
             WHERE id = ?
            """,
            (done, done, done, session.get("name"), item_id),
        )
        conn.execute(
            """
            INSERT INTO activity_updates (activity_id, user_id, user_name, message)
            VALUES (?, ?, ?, ?)
            """,
            (
                item["activity_id"],
                session.get("user_id"),
                session.get("name"),
                f"Checklist: {'marcou' if done else 'desmarcou'} o item '{item['title']}'.",
            ),
        )
        add_audit_log_conn(conn, "Marcou checklist" if done else "Desmarcou checklist", "Atividade", item["activity_id"], item["title"], "Checklist da atividade")
        conn.commit()
    return redirect(url_for("activity_detail", activity_id=item["activity_id"]) + "#checklist")


@app.route("/atividades/checklist/<int:item_id>/editar", methods=["POST"])
@login_required
def activity_checklist_edit(item_id):
    item = query_one("SELECT * FROM activity_checklist_items WHERE id = ?", (item_id,))
    if not item:
        flash("Item do checklist não encontrado.", "error")
        return redirect(url_for("activities"))
    title = request.form.get("title", "").strip()
    if not title:
        flash("O item do checklist não pode ficar vazio.", "error")
        return redirect(url_for("activity_detail", activity_id=item["activity_id"]) + "#checklist")
    execute("UPDATE activity_checklist_items SET title = ? WHERE id = ?", (title, item_id))
    add_activity_history(item["activity_id"], f"Checklist: renomeou '{item['title']}' para '{title}'.")
    audit_log("Editou item de checklist", "Atividade", item["activity_id"], title, f"Anterior: {item['title']}")
    flash("Item do checklist atualizado.", "success")
    return redirect(url_for("activity_detail", activity_id=item["activity_id"]) + "#checklist")


@app.route("/atividades/checklist/<int:item_id>/excluir", methods=["POST"])
@login_required
def activity_checklist_delete(item_id):
    item = query_one("SELECT * FROM activity_checklist_items WHERE id = ?", (item_id,))
    if not item:
        flash("Item do checklist não encontrado.", "error")
        return redirect(url_for("activities"))
    activity_id = item["activity_id"]
    execute("DELETE FROM activity_checklist_items WHERE id = ?", (item_id,))
    add_activity_history(activity_id, f"Checklist: excluiu o item '{item['title']}'.")
    audit_log("Excluiu item de checklist", "Atividade", activity_id, item["title"], "Checklist da atividade")
    flash("Item removido do checklist.", "success")
    return redirect(url_for("activity_detail", activity_id=activity_id) + "#checklist")


@app.route("/atividades/<int:activity_id>/anexos", methods=["POST"])
@login_required
def activity_upload_file(activity_id):
    activity = get_activity_or_404(activity_id)
    if not activity:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("activities"))
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        flash("Selecione um arquivo para anexar.", "error")
        return redirect(url_for("activity_detail", activity_id=activity_id))
    store_uploaded_activity_file(activity_id, uploaded)
    audit_log("Anexou arquivo", "Atividade", activity_id, activity["description"], uploaded.filename)
    flash("Arquivo anexado com sucesso.", "success")
    return redirect(url_for("activity_detail", activity_id=activity_id))


@app.route("/atividades/anexos/<int:file_id>/baixar")
@login_required
def activity_download_file(file_id):
    file_row = query_one("SELECT * FROM activity_files WHERE id = ?", (file_id,))
    if not file_row:
        flash("Arquivo não encontrado.", "error")
        return redirect(url_for("activities"))
    if row_get(file_row, "file_content") is not None:
        content = file_row["file_content"]
        if isinstance(content, memoryview):
            content = content.tobytes()
        return send_file(
            io.BytesIO(bytes(content)),
            as_attachment=True,
            download_name=file_row["original_filename"],
            mimetype=row_get(file_row, "mime_type", "application/octet-stream") or "application/octet-stream",
        )
    return send_from_directory(
        UPLOAD_DIR,
        file_row["stored_filename"],
        as_attachment=True,
        download_name=file_row["original_filename"],
    )


@app.route("/atividades/anexos/<int:file_id>/excluir", methods=["POST"])
@login_required
def activity_delete_file(file_id):
    file_row = query_one("SELECT * FROM activity_files WHERE id = ?", (file_id,))
    if not file_row:
        flash("Arquivo não encontrado.", "error")
        return redirect(url_for("activities"))
    activity_id = file_row["activity_id"]
    if file_row["stored_filename"]:
        try:
            os.remove(os.path.join(UPLOAD_DIR, file_row["stored_filename"]))
        except OSError:
            pass
    execute("DELETE FROM activity_files WHERE id = ?", (file_id,))
    audit_log("Excluiu anexo", "Atividade", activity_id, file_row["original_filename"], "Anexo removido da atividade")
    flash("Arquivo removido.", "success")
    return redirect(url_for("activity_detail", activity_id=activity_id))


@app.route("/contas-receber", methods=["GET", "POST"])
@login_required
@financial_required
def receivables():
    if request.method == "POST":
        client_id_form = int(request.form.get("client_id"))
        description = (request.form.get("description") or "").strip()
        amount = parse_money(request.form.get("amount"))
        due_date = request.form.get("due_date")
        payment_date = request.form.get("payment_date") or None
        status_form = request.form.get("status")
        payment_method = request.form.get("payment_method")
        notes = request.form.get("notes")
        installments = max(1, int(request.form.get("installments") or 1))
        receivable_group_id = str(uuid.uuid4())
        created_receivable_ids = []

        for installment_number in range(1, installments + 1):
            initial_status = status_form if installments == 1 else "A receber"
            initial_paid = amount if installments == 1 and initial_status == "Pago" else 0
            saved_payment_date = payment_date if initial_paid > 0 else None
            saved_payment_method = payment_method if initial_paid > 0 else None
            receivable_id = execute(
                """
                INSERT INTO receivables
                    (client_id, description, amount, paid_amount, due_date, payment_date, status, payment_method, notes, installment_number, installment_total, receivable_group_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    client_id_form,
                    description,
                    amount,
                    initial_paid,
                    add_months_to_date(due_date, installment_number - 1),
                    saved_payment_date,
                    initial_status,
                    saved_payment_method,
                    notes,
                    installment_number,
                    installments,
                    receivable_group_id,
                ),
            )
            if initial_paid > 0:
                execute(
                    """
                    INSERT INTO receivable_payments (receivable_id, amount, payment_date, payment_method, notes, created_by)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (receivable_id, initial_paid, saved_payment_date or date.today().isoformat(), saved_payment_method, "Baixa registrada no cadastro da cobrança", session.get("name")),
                )

        audit_log("Criou cobrança", "Cobrança", created_receivable_ids[0] if created_receivable_ids else None, description, f"Cliente ID: {client_id_form}; parcelas: {installments}; valor por parcela: {money(amount)}")
        if installments > 1:
            flash(f"Cobrança parcelada cadastrada com sucesso. Foram geradas {installments} parcelas.", "success")
        else:
            flash("Cobrança cadastrada com sucesso.", "success")
        return redirect(url_for("receivables"))

    month = request.args.get("month") or current_month()
    status = request.args.get("status", "").strip()
    client_id = request.args.get("client_id", "")
    installment_filter = request.args.get("installment_filter", "").strip()
    sort = request.args.get("sort", "due_date")
    start, end = month_range(month)
    # Regra financeira: cobranças não pagas vencidas continuam aparecendo nos meses seguintes
    # até serem baixadas como pagas ou canceladas. Assim o mês atual também mostra pendências antigas.
    filters = [
        """
        (
            EXISTS (
                SELECT 1 FROM receivable_payments p
                WHERE p.receivable_id = r.id
                  AND p.payment_date >= ?
                  AND p.payment_date < ?
            )
            OR (r.status != 'Cancelado' AND (r.amount - COALESCE(r.paid_amount, 0)) > 0 AND r.due_date < ?)
            OR (r.status = 'Cancelado' AND r.due_date >= ? AND r.due_date < ?)
        )
        """
    ]
    params = [start, end, end, start, end]
    if client_id:
        filters.append("r.client_id = ?")
        params.append(client_id)
    if installment_filter == "parcelado":
        filters.append("COALESCE(r.installment_total, 1) > 1")
    elif installment_filter == "avista":
        filters.append("COALESCE(r.installment_total, 1) <= 1")

    sort_options = {
        "due_date": "r.due_date ASC, c.name ASC, r.installment_number ASC",
        "client_az": "c.name COLLATE NOCASE ASC, r.due_date ASC, r.installment_number ASC",
        "client_za": "c.name COLLATE NOCASE DESC, r.due_date ASC, r.installment_number ASC",
        "status": "r.status COLLATE NOCASE ASC, c.name COLLATE NOCASE ASC, r.due_date ASC",
        "valor_maior": "r.amount DESC, c.name COLLATE NOCASE ASC",
        "valor_menor": "r.amount ASC, c.name COLLATE NOCASE ASC",
    }
    order_by = sort_options.get(sort, sort_options["due_date"])
    rows = query_all(
        f"""
        SELECT r.*, c.name AS client_name
        FROM receivables r
        JOIN clients c ON c.id = r.client_id
        WHERE {' AND '.join(filters)}
        ORDER BY {order_by}
        """,
        params,
    )
    if status:
        rows = [r for r in rows if computed_receivable_status(r) == status]

    payment_amounts = {}
    if rows:
        ids = [r["id"] for r in rows]
        placeholders = ",".join("?" for _ in ids)
        payment_rows = query_all(
            f"""
            SELECT receivable_id, COALESCE(SUM(amount), 0) AS total
            FROM receivable_payments
            WHERE payment_date >= ? AND payment_date < ?
              AND receivable_id IN ({placeholders})
            GROUP BY receivable_id
            """,
            [start, end, *ids],
        )
        payment_amounts = {p["receivable_id"]: float(p["total"] or 0) for p in payment_rows}
    totals = {"to_receive": 0, "received": 0, "overdue": 0}
    inadimplentes = {}
    forecast = {}
    for r in rows:
        st = computed_receivable_status(r)
        remaining = receivable_remaining_amount(r)
        totals["received"] += payment_amounts.get(r["id"], 0)
        if st != "Cancelado" and r["due_date"] and r["due_date"] < end and remaining > 0:
            totals["to_receive"] += remaining
        if is_receivable_overdue(r):
            totals["overdue"] += remaining
            inadimplentes[r["client_name"]] = inadimplentes.get(r["client_name"], 0) + remaining
        forecast_month = (r["due_date"] or "")[:7]
        if st != "Cancelado" and remaining > 0 and forecast_month:
            forecast[forecast_month] = forecast.get(forecast_month, 0) + remaining
    clients_list = query_all("SELECT id, name FROM clients WHERE status != 'Inativo' ORDER BY name")
    return render_template(
        "receivables.html", receivables=rows, clients=clients_list, month=month,
        status=status, client_id=str(client_id), installment_filter=installment_filter, sort=sort,
        statuses=RECEIVABLE_STATUS,
        payment_methods=PAYMENT_METHODS, totals=totals, inadimplentes=inadimplentes,
        forecast=dict(sorted(forecast.items())), today=date.today().isoformat(),
        computed_receivable_status=computed_receivable_status,
        receivable_paid_amount=receivable_paid_amount,
        receivable_remaining_amount=receivable_remaining_amount,
        can_receive_payment=can_receive_payment,
        can_cancel_receivable=can_cancel_receivable,
        has_receivable_payment=has_receivable_payment
    )


@app.route("/contas-receber/<int:receivable_id>/pagar", methods=["POST"])
@login_required
@financial_required
def mark_paid(receivable_id):
    receivable = query_one("SELECT * FROM receivables WHERE id = ?", (receivable_id,))
    if not receivable:
        flash("Cobrança não encontrada.", "error")
        return redirect(url_for("receivables"))
    if is_receivable_cancelled(receivable):
        flash("Esta parcela está cancelada e não pode ser baixada. Edite ou recrie a cobrança se necessário.", "error")
        return redirect(request.referrer or url_for("receivables"))
    if is_receivable_paid(receivable):
        flash("Esta parcela já está totalmente paga.", "warning")
        return redirect(request.referrer or url_for("receivables"))

    payment_method = request.form.get("payment_method") or "Pix"
    payment_date = request.form.get("payment_date") or date.today().isoformat()
    payment_amount_raw = request.form.get("payment_amount")
    remaining_before = receivable_remaining_amount(receivable)
    payment_amount = parse_money(payment_amount_raw) if payment_amount_raw else remaining_before

    if payment_amount <= 0:
        flash("Informe um valor de baixa maior que zero.", "error")
        return redirect(request.referrer or url_for("receivables"))
    if payment_amount > remaining_before:
        payment_amount = remaining_before

    new_paid_amount = receivable_paid_amount(receivable) + payment_amount
    final_status = "Pago" if new_paid_amount >= receivable_base_amount(receivable) - 0.009 else "Parcial"

    with get_db() as conn:
        conn.execute(
            """
            INSERT INTO receivable_payments (receivable_id, amount, payment_date, payment_method, notes, created_by)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (receivable_id, payment_amount, payment_date, payment_method, request.form.get("payment_notes"), session.get("name")),
        )
        conn.execute(
            """
            UPDATE receivables
               SET paid_amount = ?, status = ?, payment_date = ?, payment_method = ?, updated_at = CURRENT_TIMESTAMP
             WHERE id = ?
            """,
            (new_paid_amount, final_status, payment_date, payment_method, receivable_id),
        )
        add_audit_log_conn(conn, "Baixou cobrança", "Cobrança", receivable_id, receivable["description"], f"Valor: {money(payment_amount)}; status: {final_status}; forma: {payment_method}")
        conn.commit()

    if final_status == "Pago":
        flash("Parcela baixada integralmente.", "success")
    else:
        flash(f"Baixa parcial registrada. Saldo restante: {money(receivable_base_amount(receivable) - new_paid_amount)}.", "success")
    return redirect(request.referrer or url_for("receivables"))


@app.route("/contas-receber/<int:receivable_id>/desfazer-baixa", methods=["POST"])
@login_required
@admin_required
@financial_required
def receivable_undo_payment(receivable_id):
    receivable = query_one("SELECT * FROM receivables WHERE id = ?", (receivable_id,))
    if not receivable:
        flash("Cobrança não encontrada.", "error")
        return redirect(url_for("receivables"))
    if is_receivable_cancelled(receivable):
        flash("Parcela cancelada não possui baixa ativa para desfazer.", "warning")
        return redirect(request.referrer or url_for("receivables"))
    if not has_receivable_payment(receivable):
        flash("Esta parcela ainda não possui baixa registrada.", "warning")
        return redirect(request.referrer or url_for("receivables"))

    with get_db() as conn:
        conn.execute("DELETE FROM receivable_payments WHERE receivable_id = ?", (receivable_id,))
        conn.execute(
            """
            UPDATE receivables
               SET paid_amount = 0,
                   status = 'A receber',
                   payment_date = NULL,
                   payment_method = NULL,
                   updated_at = CURRENT_TIMESTAMP
             WHERE id = ?
            """,
            (receivable_id,),
        )
        add_audit_log_conn(conn, "Desfez baixa", "Cobrança", receivable_id, receivable["description"], f"Valor que estava baixado: {money(receivable_paid_amount(receivable))}")
        conn.commit()
    flash("Baixa desfeita. A parcela voltou para o saldo em aberto.", "success")
    return redirect(request.referrer or url_for("receivables"))


@app.route("/contas-receber/<int:receivable_id>")
@login_required
@financial_required
def receivable_detail(receivable_id):
    receivable = query_one(
        """
        SELECT r.*, c.name AS client_name
        FROM receivables r
        JOIN clients c ON c.id = r.client_id
        WHERE r.id = ?
        """,
        (receivable_id,),
    )
    if not receivable:
        flash("Cobrança não encontrada.", "error")
        return redirect(url_for("receivables"))

    if receivable["receivable_group_id"]:
        installments = query_all(
            """
            SELECT r.*, c.name AS client_name
            FROM receivables r
            JOIN clients c ON c.id = r.client_id
            WHERE r.receivable_group_id = ?
            ORDER BY r.installment_number ASC, r.due_date ASC
            """,
            (receivable["receivable_group_id"],),
        )
    else:
        installments = [receivable]

    payment_filter_sql = "r.receivable_group_id = ?" if receivable["receivable_group_id"] else "r.id = ?"
    payment_filter_value = receivable["receivable_group_id"] if receivable["receivable_group_id"] else receivable_id
    payments = query_all(
        f"""
        SELECT p.*, r.installment_number, r.installment_total
        FROM receivable_payments p
        JOIN receivables r ON r.id = p.receivable_id
        WHERE {payment_filter_sql}
        ORDER BY p.payment_date DESC, p.id DESC
        """,
        (payment_filter_value,),
    )

    total_amount, paid_amount, open_amount, cancelled_amount, paid_count, cancelled_count, partial_count = receivable_amounts_summary(installments)
    has_paid_installments = any(is_receivable_paid(r) for r in installments)
    has_open_installments = any(can_cancel_receivable(r) for r in installments)
    return render_template(
        "receivable_detail.html",
        receivable=receivable,
        installments=installments,
        payments=payments,
        total_amount=total_amount,
        paid_amount=paid_amount,
        open_amount=open_amount,
        cancelled_amount=cancelled_amount,
        paid_count=paid_count,
        cancelled_count=cancelled_count,
        partial_count=partial_count,
        has_paid_installments=has_paid_installments,
        has_open_installments=has_open_installments,
        payment_methods=PAYMENT_METHODS,
        today=date.today().isoformat(),
        computed_receivable_status=computed_receivable_status,
        receivable_paid_amount=receivable_paid_amount,
        receivable_remaining_amount=receivable_remaining_amount,
        can_delete_receivable=can_delete_receivable,
        can_cancel_receivable=can_cancel_receivable,
        can_receive_payment=can_receive_payment,
        has_receivable_payment=has_receivable_payment,
    )


@app.route("/contas-receber/<int:receivable_id>/editar", methods=["GET", "POST"])
@login_required
@admin_required
@financial_required
def receivable_edit(receivable_id):
    receivable = query_one(
        """
        SELECT r.*, c.name AS client_name
        FROM receivables r
        JOIN clients c ON c.id = r.client_id
        WHERE r.id = ?
        """,
        (receivable_id,),
    )
    if not receivable:
        flash("Cobrança não encontrada.", "error")
        return redirect(url_for("receivables"))

    group_id = receivable["receivable_group_id"]
    if not group_id:
        group_id = str(uuid.uuid4())
        execute("UPDATE receivables SET receivable_group_id = ? WHERE id = ?", (group_id, receivable_id))
        receivable = query_one(
            """
            SELECT r.*, c.name AS client_name
            FROM receivables r
            JOIN clients c ON c.id = r.client_id
            WHERE r.id = ?
            """,
            (receivable_id,),
        )

    installments = query_all(
        """
        SELECT r.*, c.name AS client_name
        FROM receivables r
        JOIN clients c ON c.id = r.client_id
        WHERE r.receivable_group_id = ?
        ORDER BY r.installment_number ASC, r.due_date ASC
        """,
        (group_id,),
    )
    if not installments:
        installments = [receivable]

    if request.method == "POST":
        client_id_form = int(request.form.get("client_id"))
        description = (request.form.get("description") or "").strip()
        total_amount = parse_money(request.form.get("total_amount"))
        requested_installments = max(1, int(request.form.get("installment_total") or 1))
        first_due_date = request.form.get("first_due_date") or date.today().isoformat()
        payment_method = request.form.get("payment_method") or None
        notes = request.form.get("notes")

        with get_db() as conn:
            current_rows = conn.execute(
                "SELECT * FROM receivables WHERE receivable_group_id = ? ORDER BY installment_number ASC, due_date ASC",
                (group_id,),
            ).fetchall()
            paid_rows = [r for r in current_rows if has_receivable_payment(r) or is_receivable_paid(r)]
            cancelled_rows = [r for r in current_rows if is_receivable_cancelled(r)]
            preserved_rows = paid_rows + [r for r in cancelled_rows if r["id"] not in {p["id"] for p in paid_rows}]
            preserved_numbers = {int(r["installment_number"] or 1) for r in preserved_rows}
            preserved_active_amount = sum(receivable_base_amount(r) for r in paid_rows)
            max_preserved_number = max(preserved_numbers) if preserved_numbers else 0
            installment_total = max(requested_installments, max_preserved_number, 1)

            open_numbers = [n for n in range(1, installment_total + 1) if n not in preserved_numbers]
            remaining_total = total_amount - preserved_active_amount
            if remaining_total < 0:
                remaining_total = 0

            conn.execute("""
                DELETE FROM receivables
                 WHERE receivable_group_id = ?
                   AND status != 'Cancelado'
                   AND COALESCE(paid_amount, 0) <= 0
                   AND status != 'Pago'
            """, (group_id,))
            conn.execute(
                """
                UPDATE receivables
                   SET client_id = ?, description = ?, notes = ?, installment_total = ?, updated_at=CURRENT_TIMESTAMP
                 WHERE receivable_group_id = ?
                   AND (status = 'Cancelado' OR status = 'Pago' OR COALESCE(paid_amount, 0) > 0)
                """,
                (client_id_form, description, notes, installment_total, group_id),
            )

            if open_numbers:
                total_cents = int(round(remaining_total * 100))
                base_cents = total_cents // len(open_numbers)
                remainder = total_cents % len(open_numbers)
                for index, installment_number in enumerate(open_numbers):
                    amount_cents = base_cents + (1 if index < remainder else 0)
                    installment_amount = amount_cents / 100
                    due_date = add_months_to_date(first_due_date, installment_number - 1)
                    conn.execute(
                        """
                        INSERT INTO receivables
                            (client_id, description, amount, paid_amount, due_date, payment_date, status, payment_method, notes, installment_number, installment_total, receivable_group_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            client_id_form,
                            description,
                            installment_amount,
                            0,
                            due_date,
                            None,
                            "A receber",
                            payment_method,
                            notes,
                            installment_number,
                            installment_total,
                            group_id,
                        ),
                    )
            add_audit_log_conn(conn, "Editou cobrança", "Cobrança", receivable_id, description, f"Parcelas: {installment_total}; total informado: {money(total_amount)}")
            conn.commit()

        if total_amount < preserved_active_amount:
            flash("Cobrança atualizada. Atenção: o total informado ficou menor que o valor de parcelas com baixa; essas parcelas foram preservadas.", "warning")
        elif paid_rows or cancelled_rows:
            flash("Cobrança atualizada. Parcelas com baixa/canceladas foram preservadas e as parcelas sem baixa foram recalculadas.", "success")
        else:
            flash("Cobrança atualizada e parcelas recalculadas com sucesso.", "success")

        first_row = query_one(
            "SELECT id FROM receivables WHERE receivable_group_id = ? ORDER BY installment_number ASC, due_date ASC LIMIT 1",
            (group_id,),
        )
        return redirect(url_for("receivable_detail", receivable_id=first_row["id"] if first_row else receivable_id))

    total_amount, paid_amount, open_amount, cancelled_amount, paid_count, cancelled_count, partial_count = receivable_amounts_summary(installments)
    first_due = installments[0]["due_date"] if installments else receivable["due_date"]
    clients_list = query_all("SELECT id, name FROM clients WHERE status != 'Inativo' ORDER BY name")
    return render_template(
        "receivable_edit.html",
        receivable=receivable,
        installments=installments,
        clients=clients_list,
        payment_methods=PAYMENT_METHODS,
        total_amount=total_amount,
        paid_amount=paid_amount,
        open_amount=open_amount,
        cancelled_amount=cancelled_amount,
        cancelled_count=cancelled_count,
        partial_count=partial_count,
        first_due=first_due,
        computed_receivable_status=computed_receivable_status,
        receivable_paid_amount=receivable_paid_amount,
        receivable_remaining_amount=receivable_remaining_amount,
    )


@app.route("/contas-receber/<int:receivable_id>/cancelar", methods=["POST"])
@login_required
@admin_required
@financial_required
def receivable_cancel(receivable_id):
    receivable = query_one("SELECT * FROM receivables WHERE id = ?", (receivable_id,))
    if not receivable:
        flash("Cobrança não encontrada.", "error")
        return redirect(url_for("receivables"))
    if has_receivable_payment(receivable) or is_receivable_paid(receivable):
        flash("Parcela com baixa registrada não pode ser cancelada. Desfaça a baixa ou faça um ajuste para preservar o histórico.", "error")
        return redirect(request.referrer or url_for("receivables"))
    if is_receivable_cancelled(receivable):
        flash("Esta parcela já está cancelada.", "warning")
        return redirect(request.referrer or url_for("receivables"))

    execute(
        """
        UPDATE receivables
           SET status='Cancelado', payment_date=NULL, payment_method=NULL,
               cancelled_at=CURRENT_TIMESTAMP, cancelled_by=?, updated_at=CURRENT_TIMESTAMP
         WHERE id=?
        """,
        (session.get("name"), receivable_id),
    )
    audit_log("Cancelou parcela", "Cobrança", receivable_id, receivable["description"], f"Valor: {money(receivable_base_amount(receivable))}")
    flash("Parcela cancelada. Ela permanece no histórico, mas não entra mais no saldo em aberto.", "success")
    return redirect(request.referrer or url_for("receivables"))


@app.route("/contas-receber/<int:receivable_id>/cancelar-grupo", methods=["POST"])
@login_required
@admin_required
@financial_required
def receivable_cancel_group(receivable_id):
    receivable = query_one("SELECT * FROM receivables WHERE id = ?", (receivable_id,))
    if not receivable:
        flash("Cobrança não encontrada.", "error")
        return redirect(url_for("receivables"))

    group_id = receivable["receivable_group_id"]
    if group_id:
        installments = query_all("SELECT * FROM receivables WHERE receivable_group_id = ?", (group_id,))
        cancellable_ids = [r["id"] for r in installments if can_cancel_receivable(r)]
        if not cancellable_ids:
            flash("Não há parcelas em aberto para cancelar nesta cobrança.", "warning")
            return redirect(request.referrer or url_for("receivables"))
        placeholders = ",".join("?" for _ in cancellable_ids)
        execute(
            f"""
            UPDATE receivables
               SET status='Cancelado', payment_date=NULL, payment_method=NULL,
                   cancelled_at=CURRENT_TIMESTAMP, cancelled_by=?, updated_at=CURRENT_TIMESTAMP
             WHERE id IN ({placeholders})
            """,
            [session.get("name"), *cancellable_ids],
        )
        audit_log("Cancelou cobrança em aberto", "Cobrança", receivable_id, receivable["description"], f"Parcelas canceladas: {len(cancellable_ids)}")
        flash(f"{len(cancellable_ids)} parcela(s) em aberto foram canceladas. Parcelas pagas foram preservadas.", "success")
    else:
        return receivable_cancel(receivable_id)
    return redirect(request.referrer or url_for("receivables"))


@app.route("/contas-receber/<int:receivable_id>/excluir", methods=["POST"])
@login_required
@admin_required
@financial_required
def receivable_delete(receivable_id):
    receivable = query_one("SELECT * FROM receivables WHERE id = ?", (receivable_id,))
    if not receivable:
        flash("Cobrança não encontrada.", "error")
        return redirect(url_for("receivables"))
    if not can_delete_receivable(receivable):
        flash("Parcela com baixa registrada não pode ser excluída. O histórico financeiro foi preservado.", "error")
        return redirect(request.referrer or url_for("receivables"))
    execute("DELETE FROM receivables WHERE id = ?", (receivable_id,))
    audit_log("Excluiu parcela", "Cobrança", receivable_id, receivable["description"], f"Valor: {money(receivable_base_amount(receivable))}")
    flash("Parcela excluída com sucesso.", "success")
    return redirect(request.referrer or url_for("receivables"))


@app.route("/contas-receber/<int:receivable_id>/excluir-grupo", methods=["POST"])
@login_required
@admin_required
@financial_required
def receivable_delete_group(receivable_id):
    receivable = query_one("SELECT * FROM receivables WHERE id = ?", (receivable_id,))
    if not receivable:
        flash("Cobrança não encontrada.", "error")
        return redirect(url_for("receivables"))

    if receivable["receivable_group_id"]:
        installments = query_all("SELECT * FROM receivables WHERE receivable_group_id = ?", (receivable["receivable_group_id"],))
        if any(has_receivable_payment(r) or is_receivable_paid(r) for r in installments):
            flash("Cobrança com baixa registrada não pode ser excluída. Cancele apenas as parcelas sem baixa para preservar o histórico.", "error")
            return redirect(request.referrer or url_for("receivable_detail", receivable_id=receivable_id))
        execute("DELETE FROM receivables WHERE receivable_group_id = ?", (receivable["receivable_group_id"],))
        audit_log("Excluiu cobrança inteira", "Cobrança", receivable_id, receivable["description"], f"Parcelas removidas: {len(installments)}")
        flash("Cobrança inteira excluída com sucesso.", "success")
    else:
        if has_receivable_payment(receivable) or is_receivable_paid(receivable):
            flash("Cobrança com baixa registrada não pode ser excluída. O histórico financeiro foi preservado.", "error")
            return redirect(request.referrer or url_for("receivables"))
        execute("DELETE FROM receivables WHERE id = ?", (receivable_id,))
        audit_log("Excluiu cobrança", "Cobrança", receivable_id, receivable["description"], f"Valor: {money(receivable_base_amount(receivable))}")
        flash("Cobrança excluída com sucesso.", "success")
    return redirect(url_for("receivables"))


@app.route("/relatorio-mensal")
@login_required
def monthly_report():
    month = request.args.get("month") or current_month()
    client_id = request.args.get("client_id", "")
    collaborator = request.args.get("collaborator", "").strip()
    service_type = request.args.get("service_type", "").strip()
    report = build_monthly_report(month, client_id=client_id or None, collaborator=collaborator or None, service_type=service_type or None)
    clients_list = query_all("SELECT id, name FROM clients ORDER BY name")
    most_work = sorted(report, key=lambda x: x["total_hours"], reverse=True)[:5]
    most_profitable = sorted([r for r in report if r["value_per_hour"] is not None], key=lambda x: x["value_per_hour"], reverse=True)[:5]
    low_profit = sorted([r for r in report if r["rentability"] in ("Atenção", "Pouco rentável")], key=lambda x: x["score_sort"])[:5]
    return render_template(
        "report.html", report=report, month=month, clients=clients_list, client_id=str(client_id),
        collaborator=collaborator, service_type=service_type, service_types=SERVICE_TYPES,
        most_work=most_work, most_profitable=most_profitable, low_profit=low_profit
    )


def report_rows_for_export(report, include_financial=True):
    rows = []
    for r in report:
        base = [r["name"], r["total_hours"], r["activities_count"]]
        if include_financial:
            base = [
                r["name"], r["monthly_fee"], r["total_hours"], r["activities_count"],
                "" if r["value_per_hour"] is None else round(r["value_per_hour"], 2), r["rentability"]
            ]
        rows.append(base)
    return rows


@app.route("/relatorio-mensal/exportar")
@login_required
def export_report():
    month = request.args.get("month") or current_month()
    client_id = request.args.get("client_id", "") or None
    collaborator = request.args.get("collaborator", "").strip() or None
    service_type = request.args.get("service_type", "").strip() or None
    fmt = request.args.get("format", "csv")
    report = build_monthly_report(month, client_id=client_id, collaborator=collaborator, service_type=service_type)
    include_financial = can_view_financial()
    headers = ["Cliente", "Horas gastas", "Qtd. atividades"]
    if include_financial:
        headers = [
            "Cliente", "Valor mensal cobrado", "Horas gastas", "Qtd. atividades",
            "Valor por hora", "Rentabilidade"
        ]
    rows = report_rows_for_export(report, include_financial=include_financial)

    if fmt == "xlsx":
        if Workbook is None:
            flash("Biblioteca openpyxl não instalada. Use CSV ou instale as dependências.", "error")
            return redirect(url_for("monthly_report", month=month))
        wb = Workbook()
        ws = wb.active
        ws.title = "Atividades"
        ws.append([f"Relatório de atividades - {month_label(month)}"])
        ws.append([])
        ws.append(headers)
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"relatorio_atividades_{month}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        if SimpleDocTemplate is None:
            flash("Biblioteca reportlab não instalada. Use CSV ou instale as dependências.", "error")
            return redirect(url_for("monthly_report", month=month))
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [Paragraph(f"Gestão360 Contábil - Relatório de atividades {month_label(month)}", styles["Title"]), Spacer(1, 12)]
        data = [headers] + [[str(x) for x in row] for row in rows]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
        doc.build(story)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"relatorio_atividades_{month}.pdf", mimetype="application/pdf")

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([f"Relatório de atividades - {month_label(month)}"])
    writer.writerow([])
    writer.writerow(headers)
    writer.writerows(rows)
    return Response(
        "\ufeff" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=relatorio_atividades_{month}.csv"},
    )


def date_br_text(value):
    if not value:
        return ""
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except (TypeError, ValueError):
        return str(value)


def receipt_rows_for_export(rows):
    return [
        [
            date_br_text(item["payment_date"]),
            item["client_name"],
            item["description"],
            f'{item["installment_number"]}/{item["installment_total"]}',
            date_br_text(item["due_date"]),
            item["installment_amount"],
            item["received_amount"],
            item["payment_method"],
            item["created_by"],
            item["payment_notes"],
        ]
        for item in rows
    ]


@app.route("/relatorio-recebimentos")
@login_required
@financial_required
def receipts_report():
    month = request.args.get("month") or current_month()
    client_id = request.args.get("client_id", "")
    payment_method = request.args.get("payment_method", "").strip()
    created_by = request.args.get("created_by", "").strip()

    report_data = build_receipts_report(
        month,
        client_id=client_id or None,
        payment_method=payment_method or None,
        created_by=created_by or None,
    )
    clients_list = query_all("SELECT id, name FROM clients ORDER BY name")
    start, end = month_range(month)
    responsible_rows = query_all(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(created_by), ''), 'Não informado') AS created_by
        FROM receivable_payments
        WHERE payment_date >= ? AND payment_date < ?
        ORDER BY created_by
        """,
        (start, end),
    )
    method_rows = query_all(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(payment_method), ''), 'Não informado') AS payment_method
        FROM receivable_payments
        WHERE payment_date >= ? AND payment_date < ?
        ORDER BY payment_method
        """,
        (start, end),
    )

    return render_template(
        "receipts_report.html",
        month=month,
        client_id=str(client_id),
        payment_method=payment_method,
        created_by=created_by,
        clients=clients_list,
        responsible_rows=responsible_rows,
        method_rows=method_rows,
        **report_data,
    )


@app.route("/relatorio-recebimentos/exportar")
@login_required
@financial_required
def export_receipts_report():
    month = request.args.get("month") or current_month()
    client_id = request.args.get("client_id", "") or None
    payment_method = request.args.get("payment_method", "").strip() or None
    created_by = request.args.get("created_by", "").strip() or None
    fmt = request.args.get("format", "csv")

    report_data = build_receipts_report(
        month,
        client_id=client_id,
        payment_method=payment_method,
        created_by=created_by,
    )
    headers = [
        "Data da baixa", "Cliente", "Cobrança", "Parcela", "Vencimento",
        "Valor da parcela", "Valor recebido", "Forma de pagamento", "Responsável", "Observação"
    ]
    rows = receipt_rows_for_export(report_data["rows"])
    total_received = report_data["summary"]["total_received"]

    if fmt == "xlsx":
        if Workbook is None:
            flash("Biblioteca openpyxl não instalada. Use CSV ou instale as dependências.", "error")
            return redirect(url_for("receipts_report", month=month))
        wb = Workbook()
        ws = wb.active
        ws.title = "Recebimentos"
        ws.append([f"Relatório de recebimentos - {month_label(month)}"])
        ws.append(["Total recebido", total_received])
        ws.append([])
        ws.append(headers)
        for row in rows:
            ws.append(row)
        ws.append([])
        ws.append(["TOTAL RECEBIDO", "", "", "", "", "", total_received])
        for cell in ws[4]:
            cell.font = cell.font.copy(bold=True)
        for row_number in range(5, 5 + len(rows)):
            ws.cell(row=row_number, column=6).number_format = 'R$ #,##0.00'
            ws.cell(row=row_number, column=7).number_format = 'R$ #,##0.00'
        ws["B2"].number_format = 'R$ #,##0.00'
        ws.cell(row=5 + len(rows) + 1, column=7).number_format = 'R$ #,##0.00'
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 44)
        ws.freeze_panes = "A5"
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"relatorio_recebimentos_{month}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if fmt == "pdf":
        if SimpleDocTemplate is None:
            flash("Biblioteca reportlab não instalada. Use CSV ou instale as dependências.", "error")
            return redirect(url_for("receipts_report", month=month))
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"Gestão360 Contábil - Relatório de recebimentos {month_label(month)}", styles["Title"]),
            Paragraph(f"Total recebido no período: {money(total_received)}", styles["Heading2"]),
            Spacer(1, 10),
        ]
        pdf_rows = []
        for item in report_data["rows"]:
            pdf_rows.append([
                date_br_text(item["payment_date"]),
                item["client_name"],
                item["description"],
                f'{item["installment_number"]}/{item["installment_total"]}',
                money(item["received_amount"]),
                item["payment_method"],
                item["created_by"],
            ])
        pdf_headers = ["Baixa", "Cliente", "Cobrança", "Parcela", "Recebido", "Forma", "Responsável"]
        data = [pdf_headers] + [[str(value or "") for value in row] for row in pdf_rows]
        table = Table(data, repeatRows=1, colWidths=[58, 112, 210, 42, 76, 78, 90])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#064746")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b9cbc7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f8f6")]),
        ]))
        story.append(table)
        doc.build(story)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"relatorio_recebimentos_{month}.pdf", mimetype="application/pdf")

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([f"Relatório de recebimentos - {month_label(month)}"])
    writer.writerow(["Total recebido", f"{total_received:.2f}"])
    writer.writerow([])
    writer.writerow(headers)
    writer.writerows(rows)
    writer.writerow([])
    writer.writerow(["TOTAL RECEBIDO", "", "", "", "", "", f"{total_received:.2f}"])
    return Response(
        "\ufeff" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=relatorio_recebimentos_{month}.csv"},
    )


@app.route("/alterar-senha", methods=["GET", "POST"])
@login_required
def change_password():
    user = query_one("SELECT * FROM users WHERE id = ?", (session.get("user_id"),))
    if not user:
        session.clear()
        flash("Sua sessão expirou. Faça login novamente.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not check_password_hash(user["password_hash"], current_password):
            flash("Senha atual incorreta.", "error")
            return redirect(url_for("change_password"))
        password_error = validate_password(new_password)
        if password_error:
            flash(password_error, "error")
            return redirect(url_for("change_password"))
        if new_password != confirm_password:
            flash("A confirmação da senha não confere.", "error")
            return redirect(url_for("change_password"))

        execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (generate_password_hash(new_password), session.get("user_id")),
        )
        session["using_default_password"] = False
        audit_log("Alterou própria senha", "Usuário", session.get("user_id"), session.get("name"), "Senha atualizada pelo usuário logado")
        flash("Senha alterada com sucesso.", "success")
        return redirect(url_for("dashboard"))

    return render_template("change_password.html")


@app.route("/backup")
@login_required
@admin_required
def backup_page():
    return render_template(
        "backup.html",
        backups=[] if using_remote_database() else list_backup_files(),
        remote_database=using_remote_database(),
    )


@app.route("/backup/gerar-banco", methods=["POST"])
@login_required
@admin_required
def backup_database_create():
    try:
        filename = create_database_backup()
        audit_log("Gerou backup", "Backup", None, filename, "Backup do banco de dados")
        flash(f"Backup do banco gerado: {filename}", "success")
    except Exception as exc:
        flash(f"Não foi possível gerar o backup do banco: {exc}", "error")
    return redirect(url_for("backup_page"))


@app.route("/backup/gerar-completo", methods=["POST"])
@login_required
@admin_required
def backup_full_create():
    try:
        filename = create_full_backup()
        audit_log("Gerou backup completo", "Backup", None, filename, "Banco de dados e anexos")
        flash(f"Backup completo gerado: {filename}", "success")
    except Exception as exc:
        flash(f"Não foi possível gerar o backup completo: {exc}", "error")
    return redirect(url_for("backup_page"))


@app.route("/backup/download/<path:filename>")
@login_required
@admin_required
def backup_download(filename):
    filename = secure_filename(filename)
    if not is_safe_backup_filename(filename):
        flash("Arquivo de backup inválido.", "error")
        return redirect(url_for("backup_page"))
    path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(path):
        flash("Backup não encontrado.", "error")
        return redirect(url_for("backup_page"))
    return send_from_directory(BACKUP_DIR, filename, as_attachment=True)


@app.route("/backup/excluir/<path:filename>", methods=["POST"])
@login_required
@admin_required
def backup_delete(filename):
    filename = secure_filename(filename)
    if not is_safe_backup_filename(filename):
        flash("Arquivo de backup inválido.", "error")
        return redirect(url_for("backup_page"))
    path = os.path.join(BACKUP_DIR, filename)
    if os.path.exists(path):
        os.remove(path)
        audit_log("Excluiu backup", "Backup", None, filename, "Arquivo removido da pasta de backups")
        flash("Backup excluído.", "success")
    else:
        flash("Backup não encontrado.", "error")
    return redirect(url_for("backup_page"))


@app.route("/historico")
@login_required
@admin_required
def audit_logs():
    action = request.args.get("action", "").strip()
    entity_type = request.args.get("entity_type", "").strip()
    user_name = request.args.get("user_name", "").strip()
    search = request.args.get("search", "").strip()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    filters = ["1=1"]
    params = []
    if action:
        filters.append("action = ?")
        params.append(action)
    if entity_type:
        filters.append("entity_type = ?")
        params.append(entity_type)
    if user_name:
        filters.append("user_name = ?")
        params.append(user_name)
    if search:
        filters.append("(entity_label LIKE ? OR details LIKE ? OR action LIKE ?)")
        like = f"%{search}%"
        params.extend([like, like, like])
    if start_date:
        filters.append("DATE(created_at) >= ?")
        params.append(start_date)
    if end_date:
        filters.append("DATE(created_at) <= ?")
        params.append(end_date)

    logs = query_all(
        f"""
        SELECT *
        FROM audit_logs
        WHERE {' AND '.join(filters)}
        ORDER BY created_at DESC, id DESC
        LIMIT 300
        """,
        params,
    )
    actions = query_all("SELECT DISTINCT action FROM audit_logs ORDER BY action")
    entity_types = query_all("SELECT DISTINCT entity_type FROM audit_logs ORDER BY entity_type")
    users_list = query_all("SELECT DISTINCT user_name FROM audit_logs WHERE user_name IS NOT NULL AND user_name != '' ORDER BY user_name")
    total_logs = query_one("SELECT COUNT(*) AS total FROM audit_logs")["total"]

    return render_template(
        "audit_logs.html",
        logs=logs,
        actions=actions,
        entity_types=entity_types,
        users_list=users_list,
        total_logs=total_logs,
        filters={
            "action": action,
            "entity_type": entity_type,
            "user_name": user_name,
            "search": search,
            "start_date": start_date,
            "end_date": end_date,
        },
    )


@app.route("/configuracoes", methods=["GET", "POST"])
@login_required
@admin_required
def settings():
    if request.method == "POST":
        hide = "1" if request.form.get("hide_financial_for_collab") == "on" else "0"
        allow_registration = "1" if request.form.get("allow_self_registration") == "on" else "0"
        new_registration_code = (request.form.get("new_registration_code") or "").strip()
        current_code_hash = setting_value("registration_code_hash", "")

        if allow_registration == "1" and not current_code_hash and len(new_registration_code) < 6:
            flash("Para liberar novos cadastros, defina um código do escritório com pelo menos 6 caracteres.", "error")
            return redirect(url_for("settings"))
        if new_registration_code and len(new_registration_code) < 6:
            flash("O novo código do escritório precisa ter pelo menos 6 caracteres.", "error")
            return redirect(url_for("settings"))

        with get_db() as conn:
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('hide_financial_for_collab', ?)", (hide,))
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('allow_self_registration', ?)", (allow_registration,))
            if new_registration_code:
                conn.execute(
                    "INSERT OR REPLACE INTO settings (key, value) VALUES ('registration_code_hash', ?)",
                    (generate_password_hash(new_registration_code),),
                )
            conn.commit()

        audit_log(
            "Alterou configurações",
            "Configuração",
            None,
            "Permissões e cadastro",
            f"Ocultar financeiro: {'sim' if hide == '1' else 'não'}; novos cadastros: {'sim' if allow_registration == '1' else 'não'}; código alterado: {'sim' if new_registration_code else 'não'}",
        )
        flash("Configurações salvas.", "success")
        return redirect(url_for("settings"))

    hide = setting_value("hide_financial_for_collab", "1") == "1"
    allow_registration = setting_value("allow_self_registration", "1") == "1"
    has_registration_code = bool(setting_value("registration_code_hash", ""))
    return render_template(
        "settings.html",
        hide_financial_for_collab=hide,
        allow_self_registration=allow_registration,
        has_registration_code=has_registration_code,
    )


@app.route("/usuarios", methods=["GET", "POST"])
@login_required
@admin_required
def users():
    if request.method == "POST":
        name = request.form.get("name")
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password") or ""
        role = request.form.get("role")
        password_error = validate_password(password)
        if password_error:
            flash(password_error, "error")
            return redirect(url_for("users"))
        try:
            new_user_id = execute(
                "INSERT INTO users (name, email, password_hash, role) VALUES (?, ?, ?, ?)",
                (name, email, generate_password_hash(password), role),
            )
            audit_log("Criou usuário", "Usuário", new_user_id, name, f"E-mail: {email}; perfil: {role}")
            flash("Usuário criado com sucesso.", "success")
        except (sqlite3.IntegrityError, ValueError):
            flash("Já existe um usuário com esse e-mail.", "error")
        return redirect(url_for("users"))
    rows = query_all("SELECT id, name, email, role, active, created_at FROM users ORDER BY name")
    return render_template("users.html", users=rows)


@app.route("/usuarios/<int:user_id>/perfil", methods=["POST"])
@login_required
@admin_required
def update_user_role(user_id):
    if user_id == session.get("user_id"):
        flash("Você não pode alterar o perfil da própria conta por esta tela.", "error")
        return redirect(url_for("users"))

    role = request.form.get("role")
    if role not in ("Administrador", "Colaborador"):
        flash("Perfil de usuário inválido.", "error")
        return redirect(url_for("users"))

    target_user = query_one("SELECT name, role FROM users WHERE id = ?", (user_id,))
    if not target_user:
        flash("Usuário não encontrado.", "error")
        return redirect(url_for("users"))

    execute("UPDATE users SET role = ? WHERE id = ?", (role, user_id))
    audit_log(
        "Alterou perfil de usuário",
        "Usuário",
        user_id,
        target_user["name"],
        f"Perfil anterior: {target_user['role']}; novo perfil: {role}",
    )
    flash("Perfil do usuário atualizado.", "success")
    return redirect(url_for("users"))


@app.route("/usuarios/<int:user_id>/alternar", methods=["POST"])
@login_required
@admin_required
def toggle_user(user_id):
    if user_id == session.get("user_id"):
        flash("Você não pode desativar seu próprio usuário.", "error")
    else:
        target_user = query_one("SELECT name, active FROM users WHERE id = ?", (user_id,))
        execute("UPDATE users SET active = CASE WHEN active = 1 THEN 0 ELSE 1 END WHERE id = ?", (user_id,))
        new_status = "desativado" if target_user and int(target_user["active"] or 0) == 1 else "ativado"
        audit_log("Alterou status de usuário", "Usuário", user_id, target_user["name"] if target_user else "", f"Novo status: {new_status}")
        flash("Status do usuário alterado.", "success")
    return redirect(url_for("users"))


@app.template_filter("datebr")
def datebr(value):
    if not value:
        return "-"
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return value


@app.template_filter("datetimebr")
def datetimebr(value):
    if not value:
        return "-"
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(value[:19], fmt).strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass
    return value


# A Vercel não mantém um arquivo SQLite local de forma permanente.
# O aplicativo precisa continuar importável durante o build, mesmo antes das variáveis
# de produção serem cadastradas. Quando faltar configuração, mostramos uma página clara
# em vez de derrubar o deploy com erro de inicialização.
REQUIRED_VERCEL_ENV = {
    "TURSO_DATABASE_URL": TURSO_DATABASE_URL,
    "TURSO_AUTH_TOKEN": TURSO_AUTH_TOKEN,
    "GESTAO360_SECRET": (os.environ.get("GESTAO360_SECRET") or "").strip(),
}
MISSING_VERCEL_ENV = [name for name, value in REQUIRED_VERCEL_ENV.items() if not value]


@app.before_request
def require_vercel_configuration():
    if IS_VERCEL and MISSING_VERCEL_ENV:
        return render_template(
            "vercel_setup.html",
            missing_variables=MISSING_VERCEL_ENV,
            app_name=APP_NAME,
        ), 503


# Inicializa o banco local normalmente e o banco remoto apenas quando a configuração
# da Vercel estiver completa.
if not (IS_VERCEL and MISSING_VERCEL_ENV):
    init_db()


if __name__ == "__main__":
    host = os.environ.get("GESTAO360_HOST", "127.0.0.1")
    port = int(os.environ.get("GESTAO360_PORT", "5000"))
    debug = os.environ.get("GESTAO360_DEBUG", "1") == "1"
    app.run(debug=debug, host=host, port=port)

